from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook


def test_schemas_infer_successful_workbook(client: TestClient) -> None:
    workbook = _build_known_manifest_workbook()

    response = client.post(
        "/schemas/infer",
        files={
            "workbook": (
                "manifest.xlsx",
                workbook,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
        data={"persist_registry": "true", "reuse_registry": "true"},
    )

    assert response.status_code == 200
    body = response.json()
    assert len(body["canonical_types_found"]) == 7
    assert body["confidence_summary"] == {
        "high": 7,
        "medium": 0,
        "low": 0,
        "none": 0,
    }
    assert body["fallback_flags"] == []
    assert body["persisted_to_registry"] is True
    assert body["reused_from_registry"] is False

    registry_response = client.get("/schemas/registry")
    assert registry_response.status_code == 200
    registry = registry_response.json()
    assert len(registry) == 1
    assert registry[0]["fingerprint"] == body["fingerprint"]
    assert len(registry[0]["mappings"]) == 7


def test_schemas_infer_partial_workbook_returns_fallbacks(client: TestClient) -> None:
    workbook = _build_partial_manifest_workbook()

    response = client.post(
        "/schemas/infer",
        files={
            "workbook": (
                "partial.xlsx",
                workbook,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
        data={"persist_registry": "false", "reuse_registry": "true"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["canonical_types_found"] == ["topic_ordering"]
    assert "missing_chapter_toc" in body["fallback_flags"]
    assert "missing_topic_ordering" not in body["fallback_flags"]
    assert body["confidence_summary"]["none"] == 6


def test_schemas_infer_reuses_registry_hit(client: TestClient) -> None:
    workbook = _build_partial_manifest_workbook()

    first_response = client.post(
        "/schemas/infer",
        files={
            "workbook": (
                "partial.xlsx",
                workbook,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
        data={"persist_registry": "true", "reuse_registry": "true"},
    )
    assert first_response.status_code == 200

    second_response = client.post(
        "/schemas/infer",
        files={
            "workbook": (
                "partial.xlsx",
                _build_partial_manifest_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
        data={"persist_registry": "false", "reuse_registry": "true"},
    )

    assert second_response.status_code == 200
    body = second_response.json()
    assert body["reused_from_registry"] is True
    assert body["persisted_to_registry"] is False
    assert body["canonical_types_found"] == ["topic_ordering"]

    registry_response = client.get("/schemas/registry")
    assert len(registry_response.json()) == 1


def test_schemas_infer_rejects_invalid_workbook(client: TestClient) -> None:
    response = client.post(
        "/schemas/infer",
        files={"workbook": ("broken.xlsx", b"not-an-xlsx", "application/octet-stream")},
        data={"persist_registry": "false", "reuse_registry": "true"},
    )

    assert response.status_code == 400
    assert response.json() == {"detail": "invalid workbook input"}


def _build_known_manifest_workbook() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_sheet(
        workbook,
        "chapter_toc",
        [
            ["chapter_id", "chapter_title", "chapter_order"],
            ["CH-01", "Introduction", 1],
            ["CH-02", "Advanced Concepts", 2],
        ],
    )
    _add_sheet(
        workbook,
        "asset_type_layout",
        [
            ["asset_type", "layout", "template"],
            ["video", "grid", "video_grid"],
        ],
    )
    _add_sheet(
        workbook,
        "topic_ordering",
        [
            ["topic_id", "topic_name", "order"],
            ["TOP-001", "Foundations", 1],
        ],
    )
    _add_sheet(
        workbook,
        "embed_registry",
        [
            ["embed_id", "embed_type", "embed_url"],
            ["EMB-01", "youtube", "https://youtube.com/watch?v=abc"],
        ],
    )
    _add_sheet(
        workbook,
        "label_map",
        [
            ["source_label", "canonical_label", "locale"],
            ["hero image", "hero_image", "en-US"],
        ],
    )
    _add_sheet(
        workbook,
        "document_url_map",
        [
            ["document_id", "document_title", "document_url"],
            ["DOC-1", "Guide", "https://cdn.example.com/guide.pdf"],
        ],
    )
    _add_sheet(
        workbook,
        "media_categories",
        [
            ["media_id", "media_type", "category"],
            ["MED-001", "video", "core"],
        ],
    )
    return _serialize_workbook(workbook)


def _build_partial_manifest_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "topic sequence"
    sheet.append(["topic name", "sort order", "chapter"])
    sheet.append(["Keyboard Basics", 1, "CH-01"])
    sheet.append(["Screen Readers", 2, "CH-01"])
    return _serialize_workbook(workbook)


def _add_sheet(workbook: Workbook, title: str, rows: list[list[object]]) -> None:
    sheet = workbook.create_sheet(title=title)
    for row in rows:
        sheet.append(row)


def _serialize_workbook(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
