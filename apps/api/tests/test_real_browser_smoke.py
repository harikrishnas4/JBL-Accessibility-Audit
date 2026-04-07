from __future__ import annotations

import json
from contextlib import contextmanager
from datetime import UTC, datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from threading import Lock, Thread
from urllib.parse import urlsplit

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from jbl_audit_api.core.dependencies import get_tier1_batch_executor
from jbl_audit_api.services.orchestration_execution import LocalBrowserWorkerBatchExecutor

FIXED_TIMESTAMP = datetime(2026, 4, 7, 15, 0, tzinfo=UTC)
SMOKE_HTML = """<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <title>Real Browser Smoke Fixture</title>
  </head>
  <body>
    <main>
      <h1>Real Browser Smoke Fixture</h1>
      <p>Minimal deterministic page for the Tier 1 browser worker smoke test.</p>
      <img src="data:image/gif;base64,R0lGODlhAQABAAAAACw=">
    </main>
  </body>
</html>
"""
PROTECTED_SMOKE_HTML = """<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <title>Authenticated Smoke Fixture</title>
  </head>
  <body data-authenticated="true">
    <main>
      <h1>Authenticated Smoke Fixture</h1>
      <p>This page is only served when the stored session state cookie is present.</p>
      <img src="data:image/gif;base64,R0lGODlhAQABAAAAACw=">
    </main>
  </body>
</html>
"""
UNAUTHORIZED_HTML = """<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <title>Unauthorized Fixture</title>
  </head>
  <body data-authenticated="false">
    <main>
      <h1>Unauthorized</h1>
      <p>Protected smoke fixture requires a session cookie.</p>
    </main>
  </body>
</html>
"""
AUTH_COOKIE_NAME = "jbl_session"
AUTH_COOKIE_VALUE = "fixture-authenticated"


class ProtectedFixtureState:
    def __init__(self) -> None:
        self.lock = Lock()
        self.authenticated_request_count = 0
        self.unauthorized_request_count = 0

    def record_authenticated(self) -> None:
        with self.lock:
            self.authenticated_request_count += 1

    def record_unauthorized(self) -> None:
        with self.lock:
            self.unauthorized_request_count += 1


class SmokeFixtureHandler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:  # noqa: N802
        parsed = urlsplit(self.path)
        if parsed.path == "/mod/page/view.php":
            payload = SMOKE_HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return

        self.send_response(404)
        self.end_headers()

    def log_message(self, format: str, *args) -> None:  # noqa: A003
        return


class ProtectedSmokeFixtureHandler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:  # noqa: N802
        parsed = urlsplit(self.path)
        if parsed.path != "/mod/page/view.php":
            self.send_response(404)
            self.end_headers()
            return

        cookie_header = self.headers.get("Cookie", "")
        expected_cookie = f"{AUTH_COOKIE_NAME}={AUTH_COOKIE_VALUE}"
        if expected_cookie in cookie_header:
            self.server.fixture_state.record_authenticated()
            payload = PROTECTED_SMOKE_HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return

        self.server.fixture_state.record_unauthorized()
        payload = UNAUTHORIZED_HTML.encode("utf-8")
        self.send_response(403)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def log_message(self, format: str, *args) -> None:  # noqa: A003
        return


@contextmanager
def smoke_fixture_server() -> str:
    server = ThreadingHTTPServer(("127.0.0.1", 0), SmokeFixtureHandler)
    thread = Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        host, port = server.server_address
        yield f"http://{host}:{port}"
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=2)


@contextmanager
def protected_smoke_fixture_server() -> tuple[str, ProtectedFixtureState]:
    server = ThreadingHTTPServer(("127.0.0.1", 0), ProtectedSmokeFixtureHandler)
    server.fixture_state = ProtectedFixtureState()
    thread = Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        host, port = server.server_address
        yield f"http://{host}:{port}", server.fixture_state
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=2)


def test_real_browser_worker_smoke_scan_persists_findings_and_report(client: TestClient) -> None:
    worker_entrypoint = client.app.state.settings.browser_worker_entrypoint
    assert worker_entrypoint.exists(), (
        "Real browser smoke test requires the built browser worker entrypoint at "
        f"{worker_entrypoint}."
    )

    client.app.dependency_overrides[get_tier1_batch_executor] = lambda: LocalBrowserWorkerBatchExecutor(
        client.app.state.settings,
    )
    try:
        with smoke_fixture_server() as base_url:
            target_url = f"{base_url}/mod/page/view.php?id=701"
            create_run_response = client.post(
                "/runs",
                json={
                    "course_url_or_name": target_url,
                    "auth_metadata": {"method": "smoke-placeholder"},
                },
            )
            assert create_run_response.status_code == 201
            run_id = create_run_response.json()["run_id"]

            auth_profile_response = client.post(
                "/auth-profiles",
                json={
                    "run_id": run_id,
                    "auth_context": {
                        "role": "learner",
                        "login_method": "manual_storage_state",
                        "captcha_bypassed_manually": False,
                        "notes": ["real browser smoke test"],
                    },
                    "session_state_path": None,
                    "validation_status": "validated",
                },
            )
            assert auth_profile_response.status_code == 201

            upsert_response = client.post(
                "/assets/upsert",
                json={
                    "run_id": run_id,
                    "crawl_snapshot": {
                        "entry_locator": target_url,
                        "started_at": FIXED_TIMESTAMP.isoformat(),
                        "completed_at": FIXED_TIMESTAMP.isoformat(),
                        "visited_locators": [target_url],
                        "excluded_locators": [],
                        "snapshot_metadata": {"visited_page_count": 1},
                    },
                    "assets": [
                        {
                            "asset_id": "smoke-web-page-701",
                            "asset_type": "web_page",
                            "source_system": "moodle",
                            "locator": target_url,
                            "scope_status": "in_scope",
                            "layer": "course_module",
                            "shared_key": "smoke:page-701",
                            "owner_team": "content",
                            "auth_context": auth_profile_response.json()["auth_context"],
                            "handling_path": "mod/page",
                            "component_fingerprint": {
                                "stable_css_selector": "main",
                                "template_id": "smoke-page",
                                "bundle_name": "view.php",
                                "controlled_dom_signature": "smoke-page-701",
                            },
                            "updated_at": FIXED_TIMESTAMP.isoformat(),
                        },
                    ],
                },
            )
            assert upsert_response.status_code == 201

        run_detail_response = client.get(f"/runs/{run_id}")
        findings_response = client.get(f"/runs/{run_id}/findings")
        defects_response = client.get("/defects", params={"run_id": run_id})

        assert run_detail_response.status_code == 200
        assert findings_response.status_code == 200
        assert defects_response.status_code == 200

        run_detail = run_detail_response.json()
        findings_body = findings_response.json()
        defects_body = defects_response.json()

        assert run_detail["status"] == "completed"
        assert run_detail["current_stage"] == "completed"
        assert run_detail["run_plan"]["status"] == "completed"

        assert findings_body["finding_count"] >= 1
        persisted_violation = next(
            finding for finding in findings_body["findings"] if finding["result_type"] == "violation"
        )
        assert persisted_violation["rule_id"] == "image-alt"
        assert persisted_violation["evidence_artifacts"]

        evidence_storage_paths = [artifact["storage_path"] for artifact in persisted_violation["evidence_artifacts"]]
        existing_evidence_files = [
            resolve_evidence_path(client.app.state.settings.evidence_root_dir, storage_path)
            for storage_path in evidence_storage_paths
        ]
        assert any(path.exists() for path in existing_evidence_files)

        assert defects_body["defect_count"] >= 1
        assert defects_body["defects"][0]["rule_id"] == "image-alt"

        assert run_detail["report_records"]
        report_path = client.app.state.settings.reports_root_dir / run_id / "wcag-audit-report.xlsx"
        assert report_path.exists()

        workbook = load_workbook(report_path)
        assert workbook.sheetnames == ["Audit Summary", "All Issues", "Component Health"]
        assert workbook["All Issues"]["A2"].value
    finally:
        client.app.dependency_overrides.pop(get_tier1_batch_executor, None)


def test_real_browser_worker_authenticated_session_smoke_scan_uses_stored_session_state(client: TestClient) -> None:
    worker_entrypoint = client.app.state.settings.browser_worker_entrypoint
    assert worker_entrypoint.exists(), (
        "Authenticated browser smoke test requires the built browser worker entrypoint at "
        f"{worker_entrypoint}."
    )

    client.app.dependency_overrides[get_tier1_batch_executor] = lambda: LocalBrowserWorkerBatchExecutor(
        client.app.state.settings,
    )
    try:
        with protected_smoke_fixture_server() as (base_url, fixture_state):
            target_url = f"{base_url}/mod/page/view.php?id=702"
            session_state_path = write_storage_state_fixture(
                client.app.state.settings.evidence_root_dir,
                base_url,
            )

            create_run_response = client.post(
                "/runs",
                json={
                    "course_url_or_name": target_url,
                    "auth_metadata": {"method": "stored-session-smoke"},
                },
            )
            assert create_run_response.status_code == 201
            run_id = create_run_response.json()["run_id"]

            auth_profile_response = client.post(
                "/auth-profiles",
                json={
                    "run_id": run_id,
                    "auth_context": {
                        "role": "learner",
                        "login_method": "manual_storage_state",
                        "captcha_bypassed_manually": False,
                        "notes": ["authenticated real browser smoke test"],
                    },
                    "session_state_path": str(session_state_path),
                    "validation_status": "validated",
                },
            )
            assert auth_profile_response.status_code == 201

            upsert_response = client.post(
                "/assets/upsert",
                json={
                    "run_id": run_id,
                    "crawl_snapshot": {
                        "entry_locator": target_url,
                        "started_at": FIXED_TIMESTAMP.isoformat(),
                        "completed_at": FIXED_TIMESTAMP.isoformat(),
                        "visited_locators": [target_url],
                        "excluded_locators": [],
                        "snapshot_metadata": {"visited_page_count": 1, "auth_mode": "stored_session_state"},
                    },
                    "assets": [
                        {
                            "asset_id": "smoke-web-page-702",
                            "asset_type": "web_page",
                            "source_system": "moodle",
                            "locator": target_url,
                            "scope_status": "in_scope",
                            "layer": "course_module",
                            "shared_key": "smoke:page-702",
                            "owner_team": "content",
                            "auth_context": auth_profile_response.json()["auth_context"],
                            "handling_path": "mod/page",
                            "component_fingerprint": {
                                "stable_css_selector": "main",
                                "template_id": "protected-smoke-page",
                                "bundle_name": "view.php",
                                "controlled_dom_signature": "smoke-page-702",
                            },
                            "updated_at": FIXED_TIMESTAMP.isoformat(),
                        },
                    ],
                },
            )
            assert upsert_response.status_code == 201

            run_detail_response = client.get(f"/runs/{run_id}")
            findings_response = client.get(f"/runs/{run_id}/findings")
            defects_response = client.get("/defects", params={"run_id": run_id})

            assert run_detail_response.status_code == 200
            assert findings_response.status_code == 200
            assert defects_response.status_code == 200

            run_detail = run_detail_response.json()
            findings_body = findings_response.json()
            defects_body = defects_response.json()

            assert fixture_state.authenticated_request_count >= 1
            assert fixture_state.unauthorized_request_count == 0

            assert run_detail["status"] == "completed"
            assert run_detail["current_stage"] == "completed"
            assert run_detail["run_plan"]["status"] == "completed"
            scan_batch = next(
                batch for batch in run_detail["run_plan"]["scan_batches"] if batch["batch_type"] == "scan_worker"
            )
            assert scan_batch["dispatcher_metadata"]["session_state_path"] == str(session_state_path)

            assert findings_body["finding_count"] >= 1
            persisted_violation = next(
                finding for finding in findings_body["findings"] if finding["result_type"] == "violation"
            )
            assert persisted_violation["rule_id"] == "image-alt"
            assert persisted_violation["evidence_artifacts"]

            evidence_files = [
                resolve_evidence_path(client.app.state.settings.evidence_root_dir, artifact["storage_path"])
                for artifact in persisted_violation["evidence_artifacts"]
            ]
            assert any(path.exists() for path in evidence_files)

            assert defects_body["defect_count"] >= 1
            assert defects_body["defects"][0]["rule_id"] == "image-alt"

            assert run_detail["report_records"]
            report_path = client.app.state.settings.reports_root_dir / run_id / "wcag-audit-report.xlsx"
            assert report_path.exists()
            workbook = load_workbook(report_path)
            assert workbook.sheetnames == ["Audit Summary", "All Issues", "Component Health"]
            assert workbook["All Issues"]["A2"].value
    finally:
        client.app.dependency_overrides.pop(get_tier1_batch_executor, None)


def write_storage_state_fixture(evidence_root_dir: Path, base_url: str) -> Path:
    host = urlsplit(base_url).hostname
    assert host is not None
    session_directory = evidence_root_dir / "session-fixtures"
    session_directory.mkdir(parents=True, exist_ok=True)
    storage_state_path = session_directory / "authenticated-storage-state.json"
    storage_state_path.write_text(
        json.dumps(
            {
                "cookies": [
                    {
                        "name": AUTH_COOKIE_NAME,
                        "value": AUTH_COOKIE_VALUE,
                        "domain": host,
                        "path": "/",
                        "expires": 2208988800,
                        "httpOnly": False,
                        "secure": False,
                        "sameSite": "Lax",
                    },
                ],
                "origins": [],
            },
        ),
        encoding="utf-8",
    )
    return storage_state_path


def resolve_evidence_path(evidence_root_dir: Path, storage_path: str) -> Path:
    parts = storage_path.split("/")
    assert parts[:2] == ["var", "evidence"]
    return evidence_root_dir.joinpath(*parts[2:])
