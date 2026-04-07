from __future__ import annotations

import uuid
from datetime import UTC, datetime

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from jbl_audit_api.db import models
from jbl_audit_api.repositories.defects import DefectRepository
from jbl_audit_api.repositories.findings import FindingRepository
from jbl_audit_api.repositories.reports import ReportRepository
from jbl_audit_api.repositories.runs import RunRepository
from jbl_audit_api.schemas.findings import EvidenceArtifactCreateRequest, RawFindingCreateRequest
from jbl_audit_api.services.findings import FindingService
from jbl_audit_api.services.normalization import NormalizationService
from jbl_audit_api.services.reporting import LocalReportStorageAdapter, ReportingService


def _create_run(client: TestClient) -> str:
    response = client.post(
        "/runs",
        json={
            "course_url_or_name": "JBL Reporting Validation Course",
            "auth_metadata": {"method": "placeholder"},
        },
    )
    return response.json()["run_id"]


def _seed_auth_profile(client: TestClient, run_id: str) -> None:
    now = datetime.now(UTC)
    with client.app.state.session_factory() as session:
        session.add(
            models.AuthProfile(
                auth_profile_id=str(uuid.uuid4()),
                run_id=run_id,
                auth_context={"role": "learner"},
                session_state_path="var/evidence/session-state.json",
                validation_status=models.AuthProfileValidationStatus.validated,
                created_at=now,
            ),
        )
        session.commit()


def _seed_classified_asset(
    client: TestClient,
    *,
    run_id: str,
    asset_id: str,
    locator: str,
    shared_key: str,
    layer: models.AssetLayer = models.AssetLayer.content,
    handling_path: models.AssetHandlingPath = models.AssetHandlingPath.automated,
    owner_team: str = "content",
    template_id: str = "lesson-header",
) -> None:
    now = datetime.now(UTC)
    with client.app.state.session_factory() as session:
        asset = models.Asset(
            run_id=run_id,
            asset_id=asset_id,
            crawl_snapshot_id=None,
            asset_type="web_page",
            source_system="moodle",
            locator=locator,
            scope_status=models.AssetScopeStatus.in_scope,
            scope_reason=None,
            layer=layer.value,
            shared_key=shared_key,
            owner_team=owner_team,
            auth_context={"role": "learner"},
            handling_path=handling_path.value,
            component_fingerprint={
                "stable_css_selector": "div.lesson-header",
                "template_id": template_id,
                "bundle_name": "lesson.js",
                "controlled_dom_signature": f"sig-{asset_id}",
            },
            created_at=now,
            updated_at=now,
        )
        asset.classification_record = models.AssetClassification(
            classification_id=str(uuid.uuid4()),
            run_id=run_id,
            asset_id=asset_id,
            layer=layer,
            handling_path=handling_path,
            shared_key=shared_key,
            owner_team=owner_team,
            third_party=False,
            third_party_evidence=None,
            auth_context={"role": "learner"},
            exclusion_reason=None,
            created_at=now,
            updated_at=now,
        )
        session.add(asset)
        session.commit()


def _persist_reported_finding(
    client: TestClient,
    *,
    run_id: str,
    asset_id: str,
    observed_at: datetime,
) -> None:
    with client.app.state.session_factory() as session:
        report_service = ReportingService(
            ReportRepository(session),
            LocalReportStorageAdapter(client.app.state.settings.reports_root_dir),
        )
        normalization_service = NormalizationService(
            DefectRepository(session),
            RunRepository(session),
            report_service=report_service,
        )
        service = FindingService(
            FindingRepository(session),
            RunRepository(session),
            normalization_service,
        )
        service.persist_scan_results(
            run_id,
            asset_id,
            [
                RawFindingCreateRequest(
                    result_type=models.RawFindingResultType.violation,
                    rule_id="button-name",
                    wcag_sc="4.1.2",
                    resolution_state="new",
                    severity="critical",
                    message="Buttons must have discernible text.",
                    target_fingerprint="button.next",
                    raw_payload={"origin": "automated"},
                    observed_at=observed_at,
                    evidence_artifacts=[
                        EvidenceArtifactCreateRequest(
                            artifact_type=models.EvidenceArtifactType.screenshot,
                            storage_path=f"var/evidence/{run_id}/{asset_id}/failure.png",
                            artifact_metadata={"viewport": "desktop"},
                            captured_at=observed_at,
                        ),
                    ],
                ),
            ],
        )
        session.commit()


def test_reporting_generates_excel_workbook_and_links_run_detail(client: TestClient) -> None:
    run_id = _create_run(client)
    _seed_auth_profile(client, run_id)
    _seed_classified_asset(
        client,
        run_id=run_id,
        asset_id="asset-1",
        locator="https://example.com/mod/page/view.php?id=101",
        shared_key="shared:lesson-header",
    )
    _seed_classified_asset(
        client,
        run_id=run_id,
        asset_id="asset-2",
        locator="https://example.com/mod/page/view.php?id=102",
        shared_key="shared:lesson-header",
    )

    observed_at = datetime(2026, 4, 6, 9, 30, tzinfo=UTC)
    _persist_reported_finding(client, run_id=run_id, asset_id="asset-1", observed_at=observed_at)
    _persist_reported_finding(client, run_id=run_id, asset_id="asset-2", observed_at=observed_at)

    run_response = client.get(f"/runs/{run_id}")

    assert run_response.status_code == 200
    report_records = run_response.json()["report_records"]
    assert len(report_records) == 1
    assert report_records[0]["report_type"] == "excel_export"
    assert report_records[0]["report_uri"] == f"var/reports/{run_id}/wcag-audit-report.xlsx"

    report_path = client.app.state.settings.reports_root_dir / run_id / "wcag-audit-report.xlsx"
    assert report_path.exists()

    workbook = load_workbook(report_path)
    assert workbook.sheetnames == ["Audit Summary", "All Issues", "Component Health"]

    summary_sheet = workbook["Audit Summary"]
    assert [summary_sheet["A1"].value, summary_sheet["B1"].value] == ["Field", "Value"]
    summary_rows = {
        summary_sheet[f"A{row_index}"].value: summary_sheet[f"B{row_index}"].value
        for row_index in range(2, summary_sheet.max_row + 1)
    }
    assert summary_rows["course"] == "JBL Reporting Validation Course"
    assert summary_rows["total_assets"] == 2
    assert summary_rows["defect_count"] == 1
    assert summary_rows["shared_assets"] == 2
    assert summary_rows["unique_assets"] == 0
    assert summary_rows["pre_check_status"] == "validated"
    assert summary_rows["scan_date"] == "2026-04-06"

    issues_sheet = workbook["All Issues"]
    issue_headers = [issues_sheet.cell(row=1, column=index).value for index in range(1, 15)]
    assert issue_headers == [
        "issue id",
        "WCAG SC",
        "priority",
        "layer",
        "owner_team",
        "asset URL/locator",
        "asset_type",
        "defect description",
        "HOW TO FIX",
        "evidence path",
        "impacted_asset_count",
        "provider_name",
        "third_party_evidence_status",
        "third_party_evidence_type",
    ]
    assert issues_sheet.max_row == 3
    assert str(issues_sheet["A2"].value).startswith("CS-")
    assert issues_sheet["B2"].value == "4.1.2"
    assert issues_sheet["C2"].value == "P1"
    assert issues_sheet["F2"].value == "https://example.com/mod/page/view.php?id=101"
    assert issues_sheet["G2"].value == "web_page"
    assert issues_sheet["H2"].value == "Buttons must have discernible text."
    assert issues_sheet["I2"].value.startswith("HOW TO FIX guidance placeholder")
    assert issues_sheet["J2"].value == f"var/evidence/{run_id}/asset-1/failure.png"
    assert issues_sheet["K2"].value == 2
    assert issues_sheet["L2"].value in {"", None}
    assert issues_sheet["M2"].value in {"", None}
    assert issues_sheet["N2"].value in {"", None}

    component_sheet = workbook["Component Health"]
    component_headers = [component_sheet.cell(row=1, column=index).value for index in range(1, 13)]
    assert component_headers == [
        "shared_key",
        "component label",
        "impacted_asset_count",
        "P1",
        "P2",
        "P3",
        "P4",
        "worst severity",
        "owner_team",
        "provider_name",
        "third_party_evidence_status",
        "third_party_evidence_type",
    ]
    assert component_sheet["A2"].value == "shared:lesson-header"
    assert component_sheet["B2"].value == "lesson-header"
    assert component_sheet["C2"].value == 2
    assert component_sheet["D2"].value == 1
    assert component_sheet["E2"].value == 0
    assert component_sheet["H2"].value == "critical"
    assert component_sheet["I2"].value == "content"
    assert component_sheet["J2"].value in {"", None}
    assert component_sheet["K2"].value in {"", None}
    assert component_sheet["L2"].value in {"", None}


def test_reporting_reuses_single_report_record_on_regeneration(client: TestClient) -> None:
    run_id = _create_run(client)
    _seed_classified_asset(
        client,
        run_id=run_id,
        asset_id="asset-1",
        locator="https://example.com/mod/page/view.php?id=201",
        shared_key="shared:lesson-header",
    )

    observed_at = datetime(2026, 4, 6, 11, 0, tzinfo=UTC)
    _persist_reported_finding(client, run_id=run_id, asset_id="asset-1", observed_at=observed_at)
    _persist_reported_finding(client, run_id=run_id, asset_id="asset-1", observed_at=observed_at)

    with client.app.state.session_factory() as session:
        records = session.scalars(
            select(models.ReportRecord).where(models.ReportRecord.run_id == run_id),
        ).all()

    assert len(records) == 1
    assert records[0].report_type == "excel_export"


def test_reporting_surfaces_seeded_third_party_evidence_status(client: TestClient) -> None:
    run_id = _create_run(client)
    observed_at = datetime(2026, 4, 6, 12, 15, tzinfo=UTC)

    upsert_response = client.post(
        "/assets/upsert",
        json={
            "run_id": run_id,
            "crawl_snapshot": {
                "entry_locator": "https://courses.example.com/course/view.php?id=555",
                "started_at": observed_at.isoformat(),
                "completed_at": observed_at.isoformat(),
                "visited_locators": ["https://courses.example.com/course/view.php?id=555"],
                "excluded_locators": [],
                "snapshot_metadata": {"visited_page_count": 1},
            },
            "assets": [
                {
                    "asset_id": "asset-third-party-report-1",
                    "asset_type": "third_party_embed",
                    "source_system": "human.biodigital.com",
                    "locator": "https://human.biodigital.com/widget?be=555",
                    "scope_status": "in_scope",
                    "layer": "embedded_content",
                    "shared_key": "third_party:biodigital-widget",
                    "owner_team": None,
                    "auth_context": {"role": "learner"},
                    "handling_path": "iframe:biodigital",
                    "component_fingerprint": {
                        "stable_css_selector": "iframe#biodigital-report",
                        "template_id": "biodigital-embed",
                        "bundle_name": "widget",
                        "controlled_dom_signature": "sig-biodigital-report",
                    },
                    "updated_at": observed_at.isoformat(),
                },
            ],
        },
    )

    assert upsert_response.status_code == 201
    classification = upsert_response.json()["classifications"][0]
    assert classification["third_party_evidence"]["domain"] == "human.biodigital.com"
    assert classification["third_party_evidence"]["status"] == "cross_origin_blocked"
    assert classification["third_party_evidence"]["evidence_type"] == "VPAT_requested"

    findings_response = client.post(
        f"/runs/{run_id}/assets/asset-third-party-report-1/findings",
        json={
            "findings": [
                {
                    "result_type": "violation",
                    "rule_id": "frame-title",
                    "wcag_sc": "2.4.1",
                    "resolution_state": "new",
                    "severity": "serious",
                    "message": "Embedded frames must include an accessible title.",
                    "target_fingerprint": "iframe.biodigital-launch",
                    "raw_payload": {"origin": "automated"},
                    "observed_at": observed_at.isoformat(),
                    "evidence_artifacts": [
                        {
                            "artifact_type": "screenshot",
                            "storage_path": f"var/evidence/{run_id}/asset-third-party-report-1/failure.png",
                            "artifact_metadata": {"viewport": "desktop"},
                            "captured_at": observed_at.isoformat(),
                        },
                    ],
                },
            ],
        },
    )

    assert findings_response.status_code == 201

    report_path = client.app.state.settings.reports_root_dir / run_id / "wcag-audit-report.xlsx"
    assert report_path.exists()

    workbook = load_workbook(report_path)
    issues_sheet = workbook["All Issues"]
    assert issues_sheet.max_row == 2
    assert issues_sheet["L2"].value == "human.biodigital.com"
    assert issues_sheet["M2"].value == "cross_origin_blocked"
    assert issues_sheet["N2"].value == "VPAT_requested"

    component_sheet = workbook["Component Health"]
    assert component_sheet["J2"].value == "human.biodigital.com"
    assert component_sheet["K2"].value == "cross_origin_blocked"
    assert component_sheet["L2"].value == "VPAT_requested"


def test_reporting_reflects_shared_defect_merge_across_two_component_assets(client: TestClient) -> None:
    run_id = _create_run(client)
    _seed_auth_profile(client, run_id)
    _seed_classified_asset(
        client,
        run_id=run_id,
        asset_id="asset-component-1",
        locator="https://example.com/mod/page/view.php?id=301",
        shared_key="component:shared-widget",
        layer=models.AssetLayer.component,
        handling_path=models.AssetHandlingPath.automated_plus_manual,
        owner_team="content",
        template_id="shared-widget",
    )
    _seed_classified_asset(
        client,
        run_id=run_id,
        asset_id="asset-component-2",
        locator="https://example.com/mod/page/view.php?id=302",
        shared_key="component:shared-widget",
        layer=models.AssetLayer.component,
        handling_path=models.AssetHandlingPath.automated_plus_manual,
        owner_team="content",
        template_id="shared-widget",
    )

    observed_at = datetime(2026, 4, 6, 14, 0, tzinfo=UTC)
    _persist_reported_finding(client, run_id=run_id, asset_id="asset-component-1", observed_at=observed_at)
    _persist_reported_finding(client, run_id=run_id, asset_id="asset-component-2", observed_at=observed_at)

    defects_response = client.get("/defects", params={"run_id": run_id})
    run_response = client.get(f"/runs/{run_id}")

    assert defects_response.status_code == 200
    assert run_response.status_code == 200

    defects_body = defects_response.json()
    assert defects_body["defect_count"] == 1
    defect = defects_body["defects"][0]
    assert defect["issue_id"].startswith("CP-")
    assert defect["impacted_asset_count"] == 2
    assert defect["shared_key"] == "component:shared-widget"
    assert {item["asset_id"] for item in defect["components"]} == {
        "asset-component-1",
        "asset-component-2",
    }

    report_records = run_response.json()["report_records"]
    assert len(report_records) == 1
    report_path = client.app.state.settings.reports_root_dir / run_id / "wcag-audit-report.xlsx"
    assert report_path.exists()

    workbook = load_workbook(report_path)
    issues_sheet = workbook["All Issues"]
    assert issues_sheet.max_row == 3
    assert str(issues_sheet["A2"].value).startswith("CP-")
    assert issues_sheet["A2"].value == issues_sheet["A3"].value
    assert issues_sheet["F2"].value == "https://example.com/mod/page/view.php?id=301"
    assert issues_sheet["F3"].value == "https://example.com/mod/page/view.php?id=302"
    assert issues_sheet["K2"].value == 2
    assert issues_sheet["K3"].value == 2

    component_sheet = workbook["Component Health"]
    assert component_sheet.max_row == 2
    assert component_sheet["A2"].value == "component:shared-widget"
    assert component_sheet["B2"].value == "shared-widget"
    assert component_sheet["C2"].value == 2
    assert component_sheet["D2"].value == 1
