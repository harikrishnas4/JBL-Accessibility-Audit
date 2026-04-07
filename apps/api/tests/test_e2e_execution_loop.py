from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import UTC, datetime
from hashlib import sha256
from html.parser import HTMLParser
from pathlib import Path, PurePosixPath
from urllib.parse import urljoin, urlsplit

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from jbl_audit_api.core.dependencies import get_tier1_batch_executor
from jbl_audit_api.db import models
from jbl_audit_api.schemas.findings import EvidenceArtifactCreateRequest, RawFindingCreateRequest
from jbl_audit_api.services.orchestration_execution import (
    Tier1AssetExecutionFailure,
    Tier1AssetExecutionSuccess,
    Tier1BatchExecutionResult,
)

FIXTURES_DIR = Path(__file__).resolve().parents[3] / "workers" / "browser" / "tests" / "fixtures"
ENTRY_URL = "https://courses.example.com/course/view.php?id=501"
FAILURE_ENTRY_URL = "https://courses.example.com/course/view.php?id=601"
FIXTURE_URLS = {
    ENTRY_URL: FIXTURES_DIR / "e2e-course-entry.html",
    FAILURE_ENTRY_URL: FIXTURES_DIR / "e2e-failure-course-entry.html",
    "https://courses.example.com/mod/page/view.php?id=510": FIXTURES_DIR / "e2e-mod-page.html",
    "https://courses.example.com/mod/page/view.php?id=610": FIXTURES_DIR / "e2e-mod-page.html",
    "https://courses.example.com/mod/page/view.php?id=611": FIXTURES_DIR / "e2e-mod-page.html",
}
FIXED_STARTED_AT = datetime(2026, 4, 7, 10, 0, tzinfo=UTC)
FIXED_COMPLETED_AT = datetime(2026, 4, 7, 10, 3, tzinfo=UTC)
FIXED_OBSERVED_AT = datetime(2026, 4, 7, 12, 0, tzinfo=UTC)


@dataclass(slots=True)
class DiscoveryNode:
    tag_name: str
    locator: str
    text_content: str
    attributes: dict[str, str]
    data_attributes: dict[str, str]
    template_id: str | None


class FixtureDiscoveryParser(HTMLParser):
    TRACKED_ATTRIBUTES = {
        "a": "href",
        "iframe": "src",
        "script": "src",
        "object": "data",
        "embed": "src",
    }

    def __init__(self, base_url: str) -> None:
        super().__init__()
        self.base_url = base_url
        self.body_template_id: str | None = None
        self.nodes: list[dict] = []
        self._open_node_stack: list[dict] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = {key: value or "" for key, value in attrs}
        if tag == "body":
            self.body_template_id = attributes.get("data-template-id")
        if tag not in self.TRACKED_ATTRIBUTES:
            return

        locator_value = attributes.get(self.TRACKED_ATTRIBUTES[tag])
        if not locator_value:
            return

        node = {
            "tag_name": tag,
            "locator": urljoin(self.base_url, locator_value),
            "attributes": {
                "id": attributes.get("id", ""),
                "class": attributes.get("class", ""),
                "role": attributes.get("role", ""),
                "type": attributes.get("type", ""),
                "data-testid": attributes.get("data-testid", ""),
                "data-template-id": attributes.get("data-template-id", ""),
            },
            "data_attributes": normalize_data_attributes(attributes),
            "template_id": attributes.get("data-template-id") or self.body_template_id,
            "text_parts": [],
        }
        self.nodes.append(node)
        if tag != "embed":
            self._open_node_stack.append(node)

    def handle_startendtag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        self.handle_starttag(tag, attrs)
        if self._open_node_stack and self._open_node_stack[-1]["tag_name"] == tag:
            self._open_node_stack.pop()

    def handle_data(self, data: str) -> None:
        if self._open_node_stack:
            self._open_node_stack[-1]["text_parts"].append(data)

    def handle_endtag(self, tag: str) -> None:
        if self._open_node_stack and self._open_node_stack[-1]["tag_name"] == tag:
            self._open_node_stack.pop()

    def as_nodes(self) -> list[DiscoveryNode]:
        parsed_nodes: list[DiscoveryNode] = []
        for node in self.nodes:
            parsed_nodes.append(
                DiscoveryNode(
                    tag_name=node["tag_name"],
                    locator=node["locator"],
                    text_content=normalize_space(" ".join(node["text_parts"])),
                    attributes=node["attributes"],
                    data_attributes=node["data_attributes"],
                    template_id=node["template_id"],
                ),
            )
        return parsed_nodes


class DeterministicExecutionLoopExecutor:
    def __init__(
        self,
        evidence_root_dir: Path,
        *,
        fixture_entry_url: str,
        fail_asset_ids: set[str] | None = None,
    ) -> None:
        self.evidence_root_dir = evidence_root_dir
        self.fixture_entry_url = fixture_entry_url
        self.fail_asset_ids = fail_asset_ids or set()

    def execute_batch(self, run_id: str, batch, *, session_state_path: str | None = None) -> Tier1BatchExecutionResult:
        asset_results: list[Tier1AssetExecutionSuccess] = []
        failures: list[Tier1AssetExecutionFailure] = []

        for asset in batch.task_contract["assets"]:
            if asset["asset_id"] in self.fail_asset_ids:
                failures.append(
                    Tier1AssetExecutionFailure(
                        asset_id=asset["asset_id"],
                        asset_type=asset["asset_type"],
                        error="simulated execution failure before findings persistence",
                        viewport="desktop",
                    ),
                )
                continue

            evidence_paths = self._write_evidence(run_id, asset["asset_id"])
            asset_results.append(
                Tier1AssetExecutionSuccess(
                    asset_id=asset["asset_id"],
                    findings=(
                        RawFindingCreateRequest(
                            result_type=models.RawFindingResultType.violation,
                            rule_id="button-name",
                            wcag_sc="4.1.2",
                            resolution_state="new",
                            severity="critical",
                            message="Buttons must have discernible text.",
                            target_fingerprint="button#next-step",
                            raw_payload={"origin": "automated", "fixture": "e2e-mod-page"},
                            observed_at=FIXED_OBSERVED_AT,
                            evidence_artifacts=[
                                EvidenceArtifactCreateRequest(
                                    artifact_type=models.EvidenceArtifactType.screenshot,
                                    storage_path=evidence_paths["screenshot"],
                                    artifact_metadata={"viewport": "desktop"},
                                    captured_at=FIXED_OBSERVED_AT,
                                ),
                                EvidenceArtifactCreateRequest(
                                    artifact_type=models.EvidenceArtifactType.trace,
                                    storage_path=evidence_paths["trace"],
                                    artifact_metadata={"viewport": "desktop"},
                                    captured_at=FIXED_OBSERVED_AT,
                                ),
                                EvidenceArtifactCreateRequest(
                                    artifact_type=models.EvidenceArtifactType.dom_snapshot_reference,
                                    storage_path=evidence_paths["dom_snapshot_reference"],
                                    artifact_metadata={"viewport": "desktop"},
                                    captured_at=FIXED_OBSERVED_AT,
                                ),
                            ],
                        ),
                        RawFindingCreateRequest(
                            result_type=models.RawFindingResultType.incomplete,
                            rule_id="color-contrast",
                            wcag_sc="1.4.3",
                            resolution_state="needs_manual_review",
                            severity="moderate",
                            message="Contrast requires manual confirmation.",
                            target_fingerprint="#lesson-shell",
                            raw_payload={"origin": "automated"},
                            observed_at=FIXED_OBSERVED_AT,
                            evidence_artifacts=[],
                        ),
                    ),
                    scan_metadata={
                        "executor": "deterministic_e2e_executor",
                        "fixture_entry_url": self.fixture_entry_url,
                        "session_state_path": session_state_path,
                        "viewports": [viewport["name"] for viewport in batch.viewport_matrix],
                    },
                ),
            )

        return Tier1BatchExecutionResult(
            asset_results=tuple(asset_results),
            failures=tuple(failures),
            summary={
                "attempted_asset_count": len(batch.task_contract["assets"]),
                "successful_asset_count": len(asset_results),
                "failed_asset_count": len(failures),
                "finding_count": sum(len(item.findings) for item in asset_results),
            },
        )

    def _write_evidence(self, run_id: str, asset_id: str) -> dict[str, str]:
        asset_directory = self.evidence_root_dir / run_id / asset_id
        asset_directory.mkdir(parents=True, exist_ok=True)
        screenshot_path = asset_directory / "failure.png"
        trace_path = asset_directory / "trace.zip"
        dom_snapshot_path = asset_directory / "dom.html"
        screenshot_path.write_bytes(b"fixture screenshot")
        trace_path.write_bytes(b"fixture trace")
        dom_snapshot_path.write_text("<html><body>fixture dom snapshot</body></html>", encoding="utf-8")
        return {
            "screenshot": f"var/evidence/{run_id}/{asset_id}/failure.png",
            "trace": f"var/evidence/{run_id}/{asset_id}/trace.zip",
            "dom_snapshot_reference": f"var/evidence/{run_id}/{asset_id}/dom.html",
        }


def test_first_execution_loop_end_to_end(client: TestClient) -> None:
    create_run_response = client.post(
        "/runs",
        json={
            "course_url_or_name": ENTRY_URL,
            "auth_metadata": {"method": "fixture-auth-placeholder"},
        },
    )
    assert create_run_response.status_code == 201
    run_body = create_run_response.json()
    run_id = run_body["run_id"]
    assert run_body["run_plan"]["status"] == "awaiting_assets"

    auth_profile_response = client.post(
        "/auth-profiles",
        json={
            "run_id": run_id,
            "auth_context": {
                "role": "learner",
                "login_method": "manual_storage_state",
                "captcha_bypassed_manually": False,
                "notes": ["fixture auth profile"],
            },
            "session_state_path": "var/evidence/session-state.json",
            "validation_status": "validated",
        },
    )
    assert auth_profile_response.status_code == 201

    client.app.dependency_overrides[get_tier1_batch_executor] = lambda: DeterministicExecutionLoopExecutor(
        client.app.state.settings.evidence_root_dir,
        fixture_entry_url=ENTRY_URL,
    )
    try:
        crawl_result = crawl_fixture_pages(
            entry_url=ENTRY_URL,
            auth_context=auth_profile_response.json()["auth_context"],
        )
        upsert_response = client.post(
            "/assets/upsert",
            json={
                "run_id": run_id,
                "crawl_snapshot": crawl_result["crawl_snapshot"],
                "assets": crawl_result["assets"],
            },
        )
    finally:
        client.app.dependency_overrides.pop(get_tier1_batch_executor, None)

    assert upsert_response.status_code == 201
    upsert_body = upsert_response.json()
    assert len(upsert_body["assets"]) == 1
    assert len(upsert_body["classifications"]) == 1
    assert upsert_body["classifications"][0]["handling_path"] == "automated"

    run_detail_response = client.get(f"/runs/{run_id}")
    findings_response = client.get(f"/runs/{run_id}/findings")
    defects_response = client.get("/defects", params={"run_id": run_id})

    assert run_detail_response.status_code == 200
    assert findings_response.status_code == 200
    assert defects_response.status_code == 200

    run_detail = run_detail_response.json()
    scan_batches = [batch for batch in run_detail["run_plan"]["scan_batches"] if batch["batch_type"] == "scan_worker"]
    assert run_detail["run_plan"]["status"] == "completed"
    assert len(scan_batches) == 1
    assert scan_batches[0]["status"] == "completed"
    assert scan_batches[0]["dispatcher_metadata"]["persisted_finding_count"] == 2
    assert scan_batches[0]["dispatcher_metadata"]["evidence_artifact_count"] == 3
    assert run_detail["report_records"]
    assert run_detail["report_records"][0]["report_uri"] == f"var/reports/{run_id}/wcag-audit-report.xlsx"

    findings_body = findings_response.json()
    assert findings_body["finding_count"] >= 1
    assert findings_body["result_counts"]["violation"] == 1
    assert findings_body["result_counts"]["incomplete"] == 1
    evidence_artifacts = [
        artifact
        for finding in findings_body["findings"]
        for artifact in finding["evidence_artifacts"]
    ]
    assert evidence_artifacts
    assert {artifact["artifact_type"] for artifact in evidence_artifacts} == {
        "screenshot",
        "trace",
        "dom_snapshot_reference",
    }

    defects_body = defects_response.json()
    assert defects_body["defect_count"] >= 1
    assert defects_body["defects"][0]["rule_id"] == "button-name"
    assert defects_body["defects"][0]["priority"] == "P1"

    report_path = client.app.state.settings.reports_root_dir / run_id / "wcag-audit-report.xlsx"
    assert report_path.exists()
    workbook = load_workbook(report_path)
    assert workbook.sheetnames == ["Audit Summary", "All Issues", "Component Health"]
    assert workbook["All Issues"]["A2"].value

    evidence_root = client.app.state.settings.evidence_root_dir / run_id
    assert (evidence_root / upsert_body["assets"][0]["asset_id"] / "failure.png").exists()
    assert (evidence_root / upsert_body["assets"][0]["asset_id"] / "trace.zip").exists()
    assert (evidence_root / upsert_body["assets"][0]["asset_id"] / "dom.html").exists()

    with client.app.state.session_factory() as session:
        manual_review_tasks = session.scalars(
            select(models.ManualReviewTask).where(models.ManualReviewTask.run_id == run_id),
        ).all()

    reasons = {task.reason for task in manual_review_tasks}
    assert "needs_manual_review" in reasons
    assert "p1_at_validation" in reasons


def test_failure_branch_preserves_persisted_outputs_and_regenerates_report(client: TestClient) -> None:
    create_run_response = client.post(
        "/runs",
        json={
            "course_url_or_name": FAILURE_ENTRY_URL,
            "auth_metadata": {"method": "fixture-auth-placeholder"},
        },
    )
    assert create_run_response.status_code == 201
    run_id = create_run_response.json()["run_id"]

    auth_profile_response = client.post(
        "/auth-profiles",
        json={
            "run_id": run_id,
            "auth_context": {
                "role": "learner",
                "login_method": "manual_storage_state",
                "captcha_bypassed_manually": False,
                "notes": ["fixture auth profile"],
            },
            "session_state_path": "var/evidence/session-state.json",
            "validation_status": "validated",
        },
    )
    assert auth_profile_response.status_code == 201

    crawl_result = crawl_fixture_pages(
        entry_url=FAILURE_ENTRY_URL,
        auth_context=auth_profile_response.json()["auth_context"],
    )
    assert len(crawl_result["assets"]) == 2
    for asset in crawl_result["assets"]:
        asset["shared_key"] = "shared:e2e-lesson-page"

    success_asset = next(asset for asset in crawl_result["assets"] if asset["locator"].endswith("id=610"))
    failed_asset = next(asset for asset in crawl_result["assets"] if asset["locator"].endswith("id=611"))

    client.app.dependency_overrides[get_tier1_batch_executor] = lambda: DeterministicExecutionLoopExecutor(
        client.app.state.settings.evidence_root_dir,
        fixture_entry_url=FAILURE_ENTRY_URL,
        fail_asset_ids={failed_asset["asset_id"]},
    )
    try:
        upsert_response = client.post(
            "/assets/upsert",
            json={
                "run_id": run_id,
                "crawl_snapshot": crawl_result["crawl_snapshot"],
                "assets": crawl_result["assets"],
            },
        )
    finally:
        client.app.dependency_overrides.pop(get_tier1_batch_executor, None)

    assert upsert_response.status_code == 201
    upsert_body = upsert_response.json()
    assert len(upsert_body["assets"]) == 2
    assert len(upsert_body["classifications"]) == 2

    run_detail_response = client.get(f"/runs/{run_id}")
    findings_response = client.get(f"/runs/{run_id}/findings")
    defects_response = client.get("/defects", params={"run_id": run_id})

    assert run_detail_response.status_code == 200
    assert findings_response.status_code == 200
    assert defects_response.status_code == 200

    run_detail = run_detail_response.json()
    assert run_detail["status"] == "failed"
    assert run_detail["current_stage"] == "failed"
    assert run_detail["run_plan"]["status"] == "failed"
    scan_batches = [batch for batch in run_detail["run_plan"]["scan_batches"] if batch["batch_type"] == "scan_worker"]
    assert len(scan_batches) == 1
    batch = scan_batches[0]
    assert batch["status"] == "failed"
    assert len(batch["asset_ids"]) == 2
    assert batch["dispatcher_metadata"]["execution_summary"] == {
        "attempted_asset_count": 2,
        "successful_asset_count": 1,
        "failed_asset_count": 1,
        "finding_count": 2,
    }
    assert batch["dispatcher_metadata"]["persisted_finding_count"] == 2
    assert batch["dispatcher_metadata"]["evidence_artifact_count"] == 3
    assert batch["dispatcher_metadata"]["failures"] == [
        {
            "asset_id": failed_asset["asset_id"],
            "asset_type": "web_page",
            "error": "simulated execution failure before findings persistence",
            "viewport": "desktop",
        },
    ]

    findings_body = findings_response.json()
    assert findings_body["finding_count"] == 2
    assert findings_body["result_counts"]["violation"] == 1
    assert findings_body["result_counts"]["incomplete"] == 1
    assert {finding["asset_id"] for finding in findings_body["findings"]} == {success_asset["asset_id"]}

    defects_body = defects_response.json()
    assert defects_body["defect_count"] == 1
    assert defects_body["defects"][0]["rule_id"] == "button-name"
    assert defects_body["defects"][0]["impacted_asset_count"] == 1
    assert {component["asset_id"] for component in defects_body["defects"][0]["components"]} == {
        success_asset["asset_id"],
    }

    assert run_detail["report_records"]
    assert run_detail["report_records"][0]["report_uri"] == f"var/reports/{run_id}/wcag-audit-report.xlsx"
    report_path = client.app.state.settings.reports_root_dir / run_id / "wcag-audit-report.xlsx"
    assert report_path.exists()
    workbook = load_workbook(report_path)
    assert workbook.sheetnames == ["Audit Summary", "All Issues", "Component Health"]
    issues_sheet = workbook["All Issues"]
    assert issues_sheet.max_row == 2
    assert issues_sheet["F2"].value == success_asset["locator"]
    assert issues_sheet["J2"].value == f"var/evidence/{run_id}/{success_asset['asset_id']}/failure.png"

    evidence_root = client.app.state.settings.evidence_root_dir / run_id
    assert (evidence_root / success_asset["asset_id"] / "failure.png").exists()
    assert (evidence_root / success_asset["asset_id"] / "trace.zip").exists()
    assert (evidence_root / success_asset["asset_id"] / "dom.html").exists()
    assert not (evidence_root / failed_asset["asset_id"]).exists()

    with client.app.state.session_factory() as session:
        manual_review_tasks = session.scalars(
            select(models.ManualReviewTask).where(models.ManualReviewTask.run_id == run_id),
        ).all()

    reasons = {task.reason for task in manual_review_tasks}
    assert "needs_manual_review" in reasons
    assert "p1_at_validation" in reasons


def test_manual_only_branch_creates_manual_tasks_without_scan_dispatch(client: TestClient) -> None:
    create_run_response = client.post(
        "/runs",
        json={
            "course_url_or_name": "https://courses.example.com/course/view.php?id=801",
            "auth_metadata": {"method": "fixture-auth-placeholder"},
        },
    )
    assert create_run_response.status_code == 201
    run_id = create_run_response.json()["run_id"]

    upsert_response = client.post(
        "/assets/upsert",
        json={
            "run_id": run_id,
            "crawl_snapshot": {
                "entry_locator": "https://courses.example.com/course/view.php?id=801",
                "started_at": FIXED_STARTED_AT.isoformat().replace("+00:00", "Z"),
                "completed_at": FIXED_COMPLETED_AT.isoformat().replace("+00:00", "Z"),
                "visited_locators": ["https://courses.example.com/course/view.php?id=801"],
                "excluded_locators": [],
                "snapshot_metadata": {"visited_page_count": 1, "asset_count": 2},
            },
            "assets": [
                {
                    "asset_id": "manual-media-1",
                    "asset_type": "media_video",
                    "source_system": "cdn-media.jblearning.com",
                    "locator": "https://cdn-media.jblearning.com/assets/lecture-1.mp4",
                    "scope_status": "in_scope",
                    "layer": "embedded_media",
                    "shared_key": "media:lecture-1",
                    "owner_team": "content",
                    "auth_context": {"role": "learner"},
                    "handling_path": "a:cdn-media",
                    "component_fingerprint": {
                        "stable_css_selector": "a#lecture-1",
                        "template_id": "resource-link",
                        "bundle_name": "lecture-1.mp4",
                        "controlled_dom_signature": "manual-media-1",
                    },
                    "updated_at": FIXED_COMPLETED_AT.isoformat().replace("+00:00", "Z"),
                },
                {
                    "asset_id": "manual-media-2",
                    "asset_type": "media_video",
                    "source_system": "cdn-media.jblearning.com",
                    "locator": "https://cdn-media.jblearning.com/assets/lecture-2.mp4",
                    "scope_status": "in_scope",
                    "layer": "embedded_media",
                    "shared_key": "media:lecture-2",
                    "owner_team": "content",
                    "auth_context": {"role": "learner"},
                    "handling_path": "a:cdn-media",
                    "component_fingerprint": {
                        "stable_css_selector": "a#lecture-2",
                        "template_id": "resource-link",
                        "bundle_name": "lecture-2.mp4",
                        "controlled_dom_signature": "manual-media-2",
                    },
                    "updated_at": FIXED_COMPLETED_AT.isoformat().replace("+00:00", "Z"),
                },
            ],
        },
    )

    assert upsert_response.status_code == 201
    upsert_body = upsert_response.json()
    assert len(upsert_body["assets"]) == 2
    assert [item["handling_path"] for item in upsert_body["classifications"]] == [
        "manual_only",
        "manual_only",
    ]
    assert {
        item["third_party_evidence"]["domain"] for item in upsert_body["classifications"]
    } == {"cdn-media.jblearning.com"}
    assert {
        item["third_party_evidence"]["status"] for item in upsert_body["classifications"]
    } == {"handling_notes_only"}

    run_detail_response = client.get(f"/runs/{run_id}")
    findings_response = client.get(f"/runs/{run_id}/findings")
    defects_response = client.get("/defects", params={"run_id": run_id})

    assert run_detail_response.status_code == 200
    assert findings_response.status_code == 200
    assert defects_response.status_code == 200

    run_detail = run_detail_response.json()
    findings_body = findings_response.json()
    defects_body = defects_response.json()

    assert run_detail["status"] == "in_progress"
    assert run_detail["current_stage"] == "orchestration"
    assert run_detail["run_plan"]["status"] == "manual_pending"
    assert run_detail["run_plan"]["scan_batch_count"] == 0
    assert run_detail["run_plan"]["manual_task_count"] == 2
    assert all(batch["batch_type"] == "manual_review_stub" for batch in run_detail["run_plan"]["scan_batches"])
    assert all(batch["status"] == "manual_pending" for batch in run_detail["run_plan"]["scan_batches"])
    assert all(
        batch["dispatcher_metadata"]["dispatch_mode"] == "manual_task_stub"
        for batch in run_detail["run_plan"]["scan_batches"]
    )
    assert all(
        batch["task_contract"]["contract_type"] == "manual_task_stub_v1"
        for batch in run_detail["run_plan"]["scan_batches"]
    )

    assert findings_body["finding_count"] == 0
    assert findings_body["result_counts"] == {
        "violation": 0,
        "pass": 0,
        "incomplete": 0,
        "inapplicable": 0,
    }
    assert defects_body["defect_count"] == 0

    assert run_detail["report_records"]
    assert run_detail["report_records"][0]["report_uri"] == f"var/reports/{run_id}/wcag-audit-report.xlsx"
    report_path = client.app.state.settings.reports_root_dir / run_id / "wcag-audit-report.xlsx"
    assert report_path.exists()

    workbook = load_workbook(report_path)
    assert workbook.sheetnames == ["Audit Summary", "All Issues", "Component Health"]
    summary_sheet = workbook["Audit Summary"]
    summary_rows = {
        summary_sheet[f"A{row_index}"].value: summary_sheet[f"B{row_index}"].value
        for row_index in range(2, summary_sheet.max_row + 1)
    }
    assert summary_rows["total_assets"] == 2
    assert summary_rows["defect_count"] == 0
    issues_sheet = workbook["All Issues"]
    assert issues_sheet.max_row == 1
    component_sheet = workbook["Component Health"]
    assert component_sheet.max_row == 3
    assert component_sheet["J2"].value == "cdn-media.jblearning.com"
    assert component_sheet["K2"].value == "handling_notes_only"
    assert component_sheet["L2"].value == "handling_notes_only"
    assert component_sheet["J3"].value == "cdn-media.jblearning.com"
    assert component_sheet["K3"].value == "handling_notes_only"
    assert component_sheet["L3"].value == "handling_notes_only"

    with client.app.state.session_factory() as session:
        manual_review_tasks = session.scalars(
            select(models.ManualReviewTask).where(models.ManualReviewTask.run_id == run_id),
        ).all()

    assert len(manual_review_tasks) == 2
    assert {task.reason for task in manual_review_tasks} == {"manual_only_asset"}
    assert {task.asset_id for task in manual_review_tasks} == {"manual-media-1", "manual-media-2"}
    assert {task.task_type.value for task in manual_review_tasks} == {"asset_review"}


def test_evidence_only_third_party_branch_skips_scan_and_preserves_evidence_linkage(client: TestClient) -> None:
    create_run_response = client.post(
        "/runs",
        json={
            "course_url_or_name": "https://courses.example.com/course/view.php?id=901",
            "auth_metadata": {"method": "fixture-auth-placeholder"},
        },
    )
    assert create_run_response.status_code == 201
    run_id = create_run_response.json()["run_id"]

    upsert_response = client.post(
        "/assets/upsert",
        json={
            "run_id": run_id,
            "crawl_snapshot": {
                "entry_locator": "https://courses.example.com/course/view.php?id=901",
                "started_at": FIXED_STARTED_AT.isoformat().replace("+00:00", "Z"),
                "completed_at": FIXED_COMPLETED_AT.isoformat().replace("+00:00", "Z"),
                "visited_locators": ["https://courses.example.com/course/view.php?id=901"],
                "excluded_locators": [],
                "snapshot_metadata": {"visited_page_count": 1, "asset_count": 2},
            },
            "assets": [
                {
                    "asset_id": "third-party-biodigital-1",
                    "asset_type": "third_party_embed",
                    "source_system": "human.biodigital.com",
                    "locator": "https://human.biodigital.com/widget?be=123",
                    "scope_status": "in_scope",
                    "layer": "embedded_content",
                    "shared_key": "third_party:biodigital-widget",
                    "owner_team": None,
                    "auth_context": {"role": "learner"},
                    "handling_path": "iframe:biodigital",
                    "component_fingerprint": {
                        "stable_css_selector": "iframe#biodigital-1",
                        "template_id": "biodigital-embed",
                        "bundle_name": "widget",
                        "controlled_dom_signature": "third-party-biodigital-1",
                    },
                    "updated_at": FIXED_COMPLETED_AT.isoformat().replace("+00:00", "Z"),
                },
                {
                    "asset_id": "third-party-biodigital-2",
                    "asset_type": "third_party_embed",
                    "source_system": "human.biodigital.com",
                    "locator": "https://human.biodigital.com/widget?be=123&instance=2",
                    "scope_status": "in_scope",
                    "layer": "embedded_content",
                    "shared_key": "third_party:biodigital-widget",
                    "owner_team": None,
                    "auth_context": {"role": "learner"},
                    "handling_path": "iframe:biodigital",
                    "component_fingerprint": {
                        "stable_css_selector": "iframe#biodigital-2",
                        "template_id": "biodigital-embed",
                        "bundle_name": "widget",
                        "controlled_dom_signature": "third-party-biodigital-2",
                    },
                    "updated_at": FIXED_COMPLETED_AT.isoformat().replace("+00:00", "Z"),
                },
            ],
        },
    )

    assert upsert_response.status_code == 201
    upsert_body = upsert_response.json()
    assert len(upsert_body["assets"]) == 2
    assert len(upsert_body["classifications"]) == 2
    assert all(item["layer"] == "third_party" for item in upsert_body["classifications"])
    assert all(item["handling_path"] == "evidence_only" for item in upsert_body["classifications"])
    assert all(item["third_party"] is True for item in upsert_body["classifications"])
    assert {
        item["third_party_evidence"]["domain"] for item in upsert_body["classifications"]
    } == {
        "human.biodigital.com",
    }
    assert {
        item["third_party_evidence"]["status"] for item in upsert_body["classifications"]
    } == {"cross_origin_blocked"}
    assert {
        item["third_party_evidence"]["evidence_type"] for item in upsert_body["classifications"]
    } == {"VPAT_requested"}

    run_detail_response = client.get(f"/runs/{run_id}")
    findings_response = client.get(f"/runs/{run_id}/findings")
    defects_response = client.get("/defects", params={"run_id": run_id})

    assert run_detail_response.status_code == 200
    assert findings_response.status_code == 200
    assert defects_response.status_code == 200

    run_detail = run_detail_response.json()
    findings_body = findings_response.json()
    defects_body = defects_response.json()

    assert run_detail["status"] == "in_progress"
    assert run_detail["current_stage"] == "orchestration"
    assert run_detail["run_plan"]["status"] == "manual_pending"
    assert run_detail["run_plan"]["scan_batch_count"] == 0
    assert run_detail["run_plan"]["manual_task_count"] == 2
    assert len(run_detail["run_plan"]["scan_batches"]) == 1
    batch = run_detail["run_plan"]["scan_batches"][0]
    assert batch["batch_type"] == "manual_review_stub"
    assert batch["status"] == "manual_pending"
    assert batch["asset_ids"] == ["third-party-biodigital-1", "third-party-biodigital-2"]
    assert batch["task_contract"]["contract_type"] == "manual_task_stub_v1"
    assert batch["dispatcher_metadata"]["dispatch_mode"] == "manual_task_stub"

    assert findings_body["finding_count"] == 0
    assert findings_body["result_counts"] == {
        "violation": 0,
        "pass": 0,
        "incomplete": 0,
        "inapplicable": 0,
    }
    assert defects_body["defect_count"] == 0

    assert run_detail["report_records"]
    assert run_detail["report_records"][0]["report_uri"] == f"var/reports/{run_id}/wcag-audit-report.xlsx"
    report_path = client.app.state.settings.reports_root_dir / run_id / "wcag-audit-report.xlsx"
    assert report_path.exists()

    workbook = load_workbook(report_path)
    assert workbook.sheetnames == ["Audit Summary", "All Issues", "Component Health"]
    issues_sheet = workbook["All Issues"]
    assert issues_sheet.max_row == 1
    component_sheet = workbook["Component Health"]
    assert component_sheet.max_row == 2
    assert component_sheet["A2"].value == "third_party:biodigital-widget"
    assert component_sheet["B2"].value == "biodigital-embed"
    assert component_sheet["C2"].value == 2
    assert component_sheet["D2"].value == 0
    assert component_sheet["E2"].value == 0
    assert component_sheet["F2"].value == 0
    assert component_sheet["G2"].value == 0
    assert component_sheet["H2"].value in {"", None}
    assert component_sheet["I2"].value == "vendor"
    assert component_sheet["J2"].value == "human.biodigital.com"
    assert component_sheet["K2"].value == "cross_origin_blocked"
    assert component_sheet["L2"].value == "VPAT_requested"

    with client.app.state.session_factory() as session:
        classifications = session.scalars(
            select(models.AssetClassification).where(models.AssetClassification.run_id == run_id),
        ).all()
        classification_domains = {
            item.third_party_evidence.domain
            for item in classifications
            if item.third_party_evidence is not None
        }
        classification_statuses = {
            item.third_party_evidence.status
            for item in classifications
            if item.third_party_evidence is not None
        }

    assert len(classifications) == 2
    assert {item.shared_key for item in classifications} == {"third_party:biodigital-widget"}
    assert classification_domains == {"human.biodigital.com"}
    assert classification_statuses == {"cross_origin_blocked"}


def crawl_fixture_pages(*, entry_url: str, auth_context: dict[str, object]) -> dict[str, object]:
    pending_urls = [entry_url]
    queued_urls = set(pending_urls)
    visited_urls: list[str] = []
    assets_by_id: dict[str, dict[str, object]] = {}
    excluded_locators: dict[str, dict[str, str]] = {}

    while pending_urls:
        current_url = pending_urls.pop(0)
        if current_url in visited_urls:
            continue

        fixture_path = FIXTURE_URLS[current_url]
        nodes = parse_fixture_nodes(fixture_path.read_text(encoding="utf-8"), current_url)
        visited_urls.append(current_url)

        for node in nodes:
            asset = classify_fixture_node(node=node, auth_context=auth_context)
            if asset is None:
                continue

            assets_by_id[asset["asset_id"]] = asset
            if asset["scope_status"] == "out_of_scope":
                excluded_locators[asset["locator"]] = {
                    "locator": asset["locator"],
                    "reason": str(asset["scope_reason"]),
                }

            locator = str(asset["locator"])
            if is_supported_module_locator(locator) and locator not in queued_urls and locator not in visited_urls:
                pending_urls.append(locator)
                queued_urls.add(locator)

    return {
        "crawl_snapshot": {
            "entry_locator": entry_url,
            "started_at": FIXED_STARTED_AT.isoformat().replace("+00:00", "Z"),
            "completed_at": FIXED_COMPLETED_AT.isoformat().replace("+00:00", "Z"),
            "visited_locators": visited_urls,
            "excluded_locators": list(excluded_locators.values()),
            "snapshot_metadata": {
                "supported_path_patterns": ["mod/page", "mod/url", "mod/quiz", "mod/lti"],
                "visited_page_count": len(visited_urls),
                "extracted_asset_count": len(assets_by_id),
                "auth_role": auth_context.get("role"),
            },
        },
        "assets": [assets_by_id[key] for key in sorted(assets_by_id)],
    }


def parse_fixture_nodes(html: str, base_url: str) -> list[DiscoveryNode]:
    parser = FixtureDiscoveryParser(base_url)
    parser.feed(html)
    return parser.as_nodes()


def classify_fixture_node(*, node: DiscoveryNode, auth_context: dict[str, object]) -> dict[str, object] | None:
    parsed = urlsplit(node.locator)
    path = parsed.path.lower()

    if "/mod/page/" in path:
        return build_asset_record(
            node=node,
            asset_type="web_page",
            source_system="moodle",
            scope_status="in_scope",
            layer="course_module",
            handling_path="mod/page",
            auth_context=auth_context,
        )

    if "/mod/url/" in path:
        return build_asset_record(
            node=node,
            asset_type="web_page",
            source_system="moodle",
            scope_status="in_scope",
            layer="course_module",
            handling_path="mod/url",
            auth_context=auth_context,
        )

    if "/mod/quiz/" in path:
        return build_asset_record(
            node=node,
            asset_type="quiz_page",
            source_system="moodle",
            scope_status="in_scope",
            layer="course_module",
            handling_path="mod/quiz",
            auth_context=auth_context,
        )

    if "/mod/lti/" in path:
        return build_asset_record(
            node=node,
            asset_type="lti_launch",
            source_system="moodle",
            scope_status="in_scope",
            layer="course_module",
            handling_path="mod/lti",
            auth_context=auth_context,
        )

    if "/mod/" in path:
        return build_asset_record(
            node=node,
            asset_type="web_page",
            source_system="moodle",
            scope_status="out_of_scope",
            scope_reason="unsupported_module_path",
            layer="course_module",
            handling_path="/".join(path.split("/")[1:3]),
            auth_context=auth_context,
        )

    return None


def build_asset_record(
    *,
    node: DiscoveryNode,
    asset_type: str,
    source_system: str,
    scope_status: str,
    layer: str,
    handling_path: str,
    auth_context: dict[str, object],
    scope_reason: str | None = None,
) -> dict[str, object]:
    fingerprint = build_component_fingerprint(node)
    record = {
        "asset_id": build_asset_id(asset_type, node.locator, fingerprint),
        "asset_type": asset_type,
        "source_system": source_system,
        "locator": node.locator,
        "scope_status": scope_status,
        "layer": layer,
        "shared_key": build_shared_key(source_system, node, fingerprint),
        "owner_team": "inventory-team",
        "auth_context": auth_context,
        "handling_path": handling_path,
        "component_fingerprint": fingerprint,
        "updated_at": FIXED_COMPLETED_AT.isoformat().replace("+00:00", "Z"),
    }
    if scope_reason is not None:
        record["scope_reason"] = scope_reason
    return record


def build_component_fingerprint(node: DiscoveryNode) -> dict[str, str | None]:
    selector = build_stable_css_selector(node)
    bundle_name = PurePosixPath(urlsplit(node.locator).path).name or None
    signature_payload = {
        "tag": node.tag_name,
        "locator_path": urlsplit(node.locator).path,
        "text": normalize_space(node.text_content)[:80],
        "attributes": node.attributes,
        "data_attributes": node.data_attributes,
        "template_id": node.template_id,
    }
    return {
        "stable_css_selector": selector,
        "template_id": node.template_id,
        "bundle_name": bundle_name,
        "controlled_dom_signature": hash_value(json.dumps(signature_payload, sort_keys=True)),
    }


def build_stable_css_selector(node: DiscoveryNode) -> str:
    if node.attributes["id"]:
        return f"{node.tag_name}#{node.attributes['id']}"
    if node.attributes["data-testid"]:
        return f'{node.tag_name}[data-testid="{node.attributes["data-testid"]}"]'
    if node.template_id:
        return f'{node.tag_name}[data-template-id="{node.template_id}"]'
    return node.tag_name


def build_asset_id(asset_type: str, locator: str, fingerprint: dict[str, str | None]) -> str:
    return hash_value(
        json.dumps(
            {
                "asset_type": asset_type,
                "locator": locator,
                "selector": fingerprint["stable_css_selector"],
                "dom_signature": fingerprint["controlled_dom_signature"],
            },
            sort_keys=True,
        ),
    )


def build_shared_key(source_system: str, node: DiscoveryNode, fingerprint: dict[str, str | None]) -> str:
    return hash_value(
        json.dumps(
            {
                "source_system": source_system,
                "locator_path": urlsplit(node.locator).path,
                "selector": fingerprint["stable_css_selector"],
                "template_id": fingerprint["template_id"],
                "bundle_name": fingerprint["bundle_name"],
                "dom_signature": fingerprint["controlled_dom_signature"],
            },
            sort_keys=True,
        ),
    )


def normalize_data_attributes(attributes: dict[str, str]) -> dict[str, str]:
    data_attributes: dict[str, str] = {}
    for key, value in attributes.items():
        if not key.startswith("data-"):
            continue
        normalized_key = key[5:].replace("-", "_")
        parts = normalized_key.split("_")
        camel_key = parts[0] + "".join(part.capitalize() for part in parts[1:])
        data_attributes[camel_key] = value
    return data_attributes


def is_supported_module_locator(locator: str) -> bool:
    path = urlsplit(locator).path.lower()
    return any(fragment in path for fragment in ("/mod/page/", "/mod/url/", "/mod/quiz/", "/mod/lti/"))


def normalize_space(value: str) -> str:
    return " ".join(value.split())


def hash_value(value: str) -> str:
    return sha256(value.encode("utf-8")).hexdigest()[:32]
