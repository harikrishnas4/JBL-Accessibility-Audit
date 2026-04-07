from __future__ import annotations

from datetime import UTC, datetime

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from jbl_audit_api.db import models
from jbl_audit_api.repositories.defects import DefectRepository
from jbl_audit_api.repositories.findings import FindingRepository
from jbl_audit_api.repositories.runs import RunRepository
from jbl_audit_api.schemas.findings import EvidenceArtifactCreateRequest, RawFindingCreateRequest
from jbl_audit_api.services.findings import FindingService
from jbl_audit_api.services.normalization import NormalizationService


def _seed_asset(client: TestClient, run_id: str, asset_id: str) -> None:
    now = datetime.now(UTC)
    with client.app.state.session_factory() as session:
        session.add(
            models.Asset(
                run_id=run_id,
                asset_id=asset_id,
                crawl_snapshot_id=None,
                asset_type="web_page",
                source_system="moodle",
                locator="https://example.com/course/page-1",
                scope_status=models.AssetScopeStatus.in_scope,
                scope_reason=None,
                layer="content",
                shared_key="shared:page-1",
                owner_team="content",
                auth_context={"role": "learner"},
                handling_path="automated",
                component_fingerprint={"stable_css_selector": "main"},
                created_at=now,
                updated_at=now,
            ),
        )
        session.commit()


def _persist_findings(client: TestClient, run_id: str, asset_id: str) -> None:
    observed_at = datetime.now(UTC)
    with client.app.state.session_factory() as session:
        service = FindingService(
            FindingRepository(session),
            RunRepository(session),
            NormalizationService(DefectRepository(session), RunRepository(session)),
        )
        service.persist_scan_results(
            run_id,
            asset_id,
            [
                RawFindingCreateRequest(
                    result_type=models.RawFindingResultType.violation,
                    rule_id="image-alt",
                    wcag_sc="1.1.1",
                    resolution_state="new",
                    severity="critical",
                    message="Images must have alternate text.",
                    target_fingerprint="img.hero",
                    raw_payload={"source": "axe"},
                    observed_at=observed_at,
                    evidence_artifacts=[
                        EvidenceArtifactCreateRequest(
                            artifact_type=models.EvidenceArtifactType.screenshot,
                            storage_path="var/evidence/run-1/asset-1/failure.png",
                            artifact_metadata={"viewport": "desktop"},
                            captured_at=observed_at,
                        ),
                        EvidenceArtifactCreateRequest(
                            artifact_type=models.EvidenceArtifactType.trace,
                            storage_path="var/evidence/run-1/asset-1/failure.zip",
                            artifact_metadata={"viewport": "desktop"},
                            captured_at=observed_at,
                        ),
                    ],
                ),
                RawFindingCreateRequest(
                    result_type=models.RawFindingResultType.pass_,
                    rule_id="page-has-heading-one",
                    wcag_sc="1.3.1",
                    resolution_state="new",
                    severity=None,
                    message="Page has a heading level one.",
                    target_fingerprint="h1",
                    raw_payload={"source": "axe"},
                    observed_at=observed_at,
                    evidence_artifacts=[
                        EvidenceArtifactCreateRequest(
                            artifact_type=models.EvidenceArtifactType.dom_snapshot_reference,
                            storage_path="var/evidence/run-1/asset-1/dom.html",
                            artifact_metadata={"content_type": "text/html"},
                            captured_at=observed_at,
                        ),
                    ],
                ),
            ],
        )
        session.commit()


def _build_ingest_payload(*, observed_at: datetime, include_scan_metadata: bool = True) -> dict:
    payload = {
        "findings": [
            {
                "result_type": "violation",
                "rule_id": "image-alt",
                "wcag_sc": "1.1.1",
                "resolution_state": "new",
                "severity": "critical",
                "message": "Images must have alternate text.",
                "target_fingerprint": "img.hero",
                "raw_payload": {"source": "axe"},
                "observed_at": observed_at.isoformat(),
                "evidence_artifacts": [
                    {
                        "artifact_type": "screenshot",
                        "storage_path": "var/evidence/run-1/asset-1/failure.png",
                        "artifact_metadata": {"viewport": "desktop"},
                        "captured_at": observed_at.isoformat(),
                    },
                    {
                        "artifact_type": "trace",
                        "storage_path": "var/evidence/run-1/asset-1/failure.zip",
                        "artifact_metadata": {"viewport": "desktop"},
                        "captured_at": observed_at.isoformat(),
                    },
                ],
            },
            {
                "result_type": "pass",
                "rule_id": "page-has-heading-one",
                "wcag_sc": "1.3.1",
                "resolution_state": "new",
                "severity": None,
                "message": "Page has a heading level one.",
                "target_fingerprint": "h1",
                "raw_payload": {"source": "axe"},
                "observed_at": observed_at.isoformat(),
                "evidence_artifacts": [
                    {
                        "artifact_type": "dom_snapshot_reference",
                        "storage_path": "var/evidence/run-1/asset-1/dom.html",
                        "artifact_metadata": {"content_type": "text/html"},
                        "captured_at": observed_at.isoformat(),
                    },
                ],
            },
        ],
    }
    if include_scan_metadata:
        payload["scan_metadata"] = {
            "viewport": "desktop",
            "adapter_version": "tier1",
        }
    return payload


def test_post_asset_findings_persists_findings_triggers_normalization_and_regenerates_report(
    client: TestClient,
) -> None:
    create_response = client.post(
        "/runs",
        json={
            "course_url_or_name": "Accessibility Sample Course",
            "auth_metadata": {"method": "placeholder"},
        },
    )
    run_id = create_response.json()["run_id"]
    _seed_asset(client, run_id, "asset-1")

    observed_at = datetime(2026, 4, 7, 9, 15, tzinfo=UTC)
    response = client.post(
        f"/runs/{run_id}/assets/asset-1/findings",
        json=_build_ingest_payload(observed_at=observed_at),
    )

    assert response.status_code == 201
    body = response.json()
    assert body == {
        "run_id": run_id,
        "asset_id": "asset-1",
        "persisted_finding_count": 2,
        "evidence_artifact_count": 3,
        "result_counts": {
            "violation": 1,
            "pass": 1,
            "incomplete": 0,
            "inapplicable": 0,
        },
        "scan_metadata": {
            "viewport": "desktop",
            "adapter_version": "tier1",
        },
    }

    findings_response = client.get(f"/runs/{run_id}/findings")
    assert findings_response.status_code == 200
    persisted_violation = next(
        item for item in findings_response.json()["findings"] if item["result_type"] == "violation"
    )
    assert persisted_violation["raw_payload"]["scan_metadata"] == {
        "viewport": "desktop",
        "adapter_version": "tier1",
    }

    run_response = client.get(f"/runs/{run_id}")
    assert run_response.status_code == 200
    assert len(run_response.json()["report_records"]) == 1
    report_path = client.app.state.settings.reports_root_dir / run_id / "wcag-audit-report.xlsx"
    assert report_path.exists()


def test_post_asset_findings_returns_404_for_unknown_run(client: TestClient) -> None:
    observed_at = datetime(2026, 4, 7, 10, 0, tzinfo=UTC)

    response = client.post(
        "/runs/does-not-exist/assets/asset-1/findings",
        json=_build_ingest_payload(observed_at=observed_at),
    )

    assert response.status_code == 404
    assert response.json() == {"detail": "run 'does-not-exist' does not exist"}


def test_post_asset_findings_returns_404_for_unknown_asset(client: TestClient) -> None:
    create_response = client.post(
        "/runs",
        json={
            "course_url_or_name": "Accessibility Sample Course",
            "auth_metadata": {"method": "placeholder"},
        },
    )
    run_id = create_response.json()["run_id"]
    observed_at = datetime(2026, 4, 7, 10, 5, tzinfo=UTC)

    response = client.post(
        f"/runs/{run_id}/assets/missing-asset/findings",
        json=_build_ingest_payload(observed_at=observed_at),
    )

    assert response.status_code == 404
    assert response.json() == {"detail": f"asset 'missing-asset' does not exist for run '{run_id}'"}


def test_post_asset_findings_supports_mixed_result_types(client: TestClient) -> None:
    create_response = client.post(
        "/runs",
        json={
            "course_url_or_name": "Accessibility Sample Course",
            "auth_metadata": {"method": "placeholder"},
        },
    )
    run_id = create_response.json()["run_id"]
    _seed_asset(client, run_id, "asset-1")
    observed_at = datetime(2026, 4, 7, 10, 30, tzinfo=UTC)

    response = client.post(
        f"/runs/{run_id}/assets/asset-1/findings",
        json={
            "findings": [
                {
                    "result_type": "violation",
                    "rule_id": "image-alt",
                    "wcag_sc": "1.1.1",
                    "resolution_state": "new",
                    "severity": "critical",
                    "message": "Images must have alternate text.",
                    "target_fingerprint": "img.hero",
                    "raw_payload": {},
                    "observed_at": observed_at.isoformat(),
                    "evidence_artifacts": [],
                },
                {
                    "result_type": "pass",
                    "rule_id": "page-has-heading-one",
                    "wcag_sc": "1.3.1",
                    "resolution_state": "new",
                    "severity": None,
                    "message": "Page has a heading level one.",
                    "target_fingerprint": "h1",
                    "raw_payload": {},
                    "observed_at": observed_at.isoformat(),
                    "evidence_artifacts": [],
                },
                {
                    "result_type": "incomplete",
                    "rule_id": "color-contrast",
                    "wcag_sc": "1.4.3",
                    "resolution_state": "new",
                    "severity": "serious",
                    "message": "Contrast could not be verified automatically.",
                    "target_fingerprint": ".button-cta",
                    "raw_payload": {},
                    "observed_at": observed_at.isoformat(),
                    "evidence_artifacts": [],
                },
                {
                    "result_type": "inapplicable",
                    "rule_id": "video-caption",
                    "wcag_sc": "1.2.2",
                    "resolution_state": "new",
                    "severity": None,
                    "message": "Video captions are not applicable.",
                    "target_fingerprint": None,
                    "raw_payload": {},
                    "observed_at": observed_at.isoformat(),
                    "evidence_artifacts": [],
                },
            ],
        },
    )

    assert response.status_code == 201
    assert response.json()["persisted_finding_count"] == 4
    assert response.json()["result_counts"] == {
        "violation": 1,
        "pass": 1,
        "incomplete": 1,
        "inapplicable": 1,
    }

    findings_response = client.get(f"/runs/{run_id}/findings")
    assert findings_response.status_code == 200
    assert findings_response.json()["finding_count"] == 4
    assert findings_response.json()["result_counts"] == {
        "violation": 1,
        "pass": 1,
        "incomplete": 1,
        "inapplicable": 1,
    }


def test_post_asset_findings_creates_manual_review_task_for_incomplete_result(client: TestClient) -> None:
    create_response = client.post(
        "/runs",
        json={
            "course_url_or_name": "Accessibility Sample Course",
            "auth_metadata": {"method": "placeholder"},
        },
    )
    run_id = create_response.json()["run_id"]
    _seed_asset(client, run_id, "asset-1")
    observed_at = datetime(2026, 4, 7, 10, 45, tzinfo=UTC)

    response = client.post(
        f"/runs/{run_id}/assets/asset-1/findings",
        json={
            "findings": [
                {
                    "result_type": "incomplete",
                    "rule_id": "color-contrast",
                    "wcag_sc": "1.4.3",
                    "resolution_state": "new",
                    "severity": "serious",
                    "message": "Contrast could not be verified automatically.",
                    "target_fingerprint": ".button-cta",
                    "raw_payload": {"origin": "automated"},
                    "observed_at": observed_at.isoformat(),
                    "evidence_artifacts": [],
                },
            ],
            "scan_metadata": {
                "viewport": "desktop",
                "adapter_version": "tier1",
            },
        },
    )

    assert response.status_code == 201
    assert response.json() == {
        "run_id": run_id,
        "asset_id": "asset-1",
        "persisted_finding_count": 1,
        "evidence_artifact_count": 0,
        "result_counts": {
            "violation": 0,
            "pass": 0,
            "incomplete": 1,
            "inapplicable": 0,
        },
        "scan_metadata": {
            "viewport": "desktop",
            "adapter_version": "tier1",
        },
    }

    findings_response = client.get(f"/runs/{run_id}/findings")
    run_response = client.get(f"/runs/{run_id}")

    assert findings_response.status_code == 200
    assert run_response.status_code == 200
    assert findings_response.json()["finding_count"] == 1
    assert findings_response.json()["result_counts"] == {
        "violation": 0,
        "pass": 0,
        "incomplete": 1,
        "inapplicable": 0,
    }
    assert len(run_response.json()["report_records"]) == 1
    report_path = client.app.state.settings.reports_root_dir / run_id / "wcag-audit-report.xlsx"
    assert report_path.exists()

    with client.app.state.session_factory() as session:
        persisted_finding = session.scalar(
            select(models.RawFinding)
            .where(models.RawFinding.run_id == run_id, models.RawFinding.asset_id == "asset-1")
        )
        manual_review_tasks = session.scalars(
            select(models.ManualReviewTask).where(models.ManualReviewTask.run_id == run_id)
        ).all()

    assert persisted_finding is not None
    assert len(manual_review_tasks) == 1
    task = manual_review_tasks[0]
    assert task.reason == "needs_manual_review"
    assert task.task_type == models.ManualReviewTaskType.finding_review
    assert task.source_state == models.FindingState.needs_manual_review
    assert task.asset_id == "asset-1"
    assert task.finding_id == persisted_finding.finding_id
    assert task.priority == models.DefectPriority.P2


def test_post_asset_findings_creates_blocked_manual_review_task_with_third_party_evidence_linkage(
    client: TestClient,
) -> None:
    create_response = client.post(
        "/runs",
        json={
            "course_url_or_name": "Blocked Third-Party Validation Course",
            "auth_metadata": {"method": "placeholder"},
        },
    )
    run_id = create_response.json()["run_id"]
    observed_at = datetime(2026, 4, 7, 11, 15, tzinfo=UTC)

    upsert_response = client.post(
        "/assets/upsert",
        json={
            "run_id": run_id,
            "crawl_snapshot": {
                "entry_locator": "https://courses.example.com/course/view.php?id=990",
                "started_at": observed_at.isoformat(),
                "completed_at": observed_at.isoformat(),
                "visited_locators": ["https://courses.example.com/course/view.php?id=990"],
                "excluded_locators": [],
                "snapshot_metadata": {"visited_page_count": 1},
            },
            "assets": [
                {
                    "asset_id": "asset-third-party-1",
                    "asset_type": "third_party_embed",
                    "source_system": "human.biodigital.com",
                    "locator": "https://human.biodigital.com/widget?be=123",
                    "scope_status": "in_scope",
                    "layer": "embedded_content",
                    "shared_key": "third_party:biodigital-widget",
                    "owner_team": None,
                    "auth_context": {"role": "learner"},
                    "handling_path": "iframe:biodigital",
                    "component_fingerprint": {
                        "stable_css_selector": "iframe#biodigital-frame",
                        "template_id": "biodigital-embed",
                        "bundle_name": "widget",
                        "controlled_dom_signature": "sig-third-party-1",
                    },
                    "updated_at": observed_at.isoformat(),
                },
            ],
        },
    )

    assert upsert_response.status_code == 201
    classification = upsert_response.json()["classifications"][0]
    assert classification["layer"] == "third_party"
    assert classification["handling_path"] == "evidence_only"
    assert classification["third_party"] is True
    assert classification["third_party_evidence"]["domain"] == "human.biodigital.com"
    assert classification["third_party_evidence"]["status"] == "cross_origin_blocked"
    assert classification["third_party_evidence"]["evidence_type"] == "VPAT_requested"

    response = client.post(
        f"/runs/{run_id}/assets/asset-third-party-1/findings",
        json={
            "findings": [
                {
                    "result_type": "violation",
                    "rule_id": "frame-title",
                    "wcag_sc": "2.4.1",
                    "resolution_state": "blocked",
                    "severity": "serious",
                    "message": "Embedded frame access was blocked.",
                    "target_fingerprint": "iframe.biodigital-launch",
                    "raw_payload": {"origin": "automated"},
                    "observed_at": observed_at.isoformat(),
                    "evidence_artifacts": [],
                },
            ],
            "scan_metadata": {
                "viewport": "desktop",
                "adapter_version": "tier1",
            },
        },
    )

    assert response.status_code == 201
    assert response.json() == {
        "run_id": run_id,
        "asset_id": "asset-third-party-1",
        "persisted_finding_count": 1,
        "evidence_artifact_count": 0,
        "result_counts": {
            "violation": 1,
            "pass": 0,
            "incomplete": 0,
            "inapplicable": 0,
        },
        "scan_metadata": {
            "viewport": "desktop",
            "adapter_version": "tier1",
        },
    }

    findings_response = client.get(f"/runs/{run_id}/findings")
    defects_response = client.get("/defects", params={"run_id": run_id})
    run_response = client.get(f"/runs/{run_id}")

    assert findings_response.status_code == 200
    assert defects_response.status_code == 200
    assert run_response.status_code == 200
    assert findings_response.json()["finding_count"] == 1
    assert findings_response.json()["findings"][0]["asset_id"] == "asset-third-party-1"
    assert defects_response.json()["defect_count"] == 0
    assert len(run_response.json()["report_records"]) == 1

    report_path = client.app.state.settings.reports_root_dir / run_id / "wcag-audit-report.xlsx"
    assert report_path.exists()
    workbook = load_workbook(report_path)
    assert workbook.sheetnames == ["Audit Summary", "All Issues", "Component Health"]
    assert workbook["All Issues"].max_row == 1

    with client.app.state.session_factory() as session:
        persisted_finding = session.scalar(
            select(models.RawFinding)
            .where(models.RawFinding.run_id == run_id, models.RawFinding.asset_id == "asset-third-party-1")
        )
        manual_review_tasks = session.scalars(
            select(models.ManualReviewTask).where(models.ManualReviewTask.run_id == run_id)
        ).all()
        persisted_classification = session.scalar(
            select(models.AssetClassification).where(
                models.AssetClassification.run_id == run_id,
                models.AssetClassification.asset_id == "asset-third-party-1",
            )
        )
        persisted_evidence_id = (
            persisted_classification.third_party_evidence.third_party_evidence_id
            if persisted_classification is not None and persisted_classification.third_party_evidence is not None
            else None
        )

    assert persisted_finding is not None
    assert persisted_classification is not None
    assert len(manual_review_tasks) == 1
    task = manual_review_tasks[0]
    assert task.reason == "blocked"
    assert task.source_state == models.FindingState.blocked
    assert task.asset_id == "asset-third-party-1"
    assert task.finding_id == persisted_finding.finding_id
    assert task.task_type == models.ManualReviewTaskType.finding_review
    assert task.task_metadata["third_party_evidence"]["third_party_evidence_id"] == persisted_evidence_id
    assert task.task_metadata["third_party_evidence"]["domain"] == "human.biodigital.com"
    assert task.task_metadata["third_party_evidence"]["status"] == "cross_origin_blocked"


def test_post_asset_findings_persists_evidence_artifacts(client: TestClient) -> None:
    create_response = client.post(
        "/runs",
        json={
            "course_url_or_name": "Accessibility Sample Course",
            "auth_metadata": {"method": "placeholder"},
        },
    )
    run_id = create_response.json()["run_id"]
    _seed_asset(client, run_id, "asset-1")
    observed_at = datetime(2026, 4, 7, 11, 0, tzinfo=UTC)

    response = client.post(
        f"/runs/{run_id}/assets/asset-1/findings",
        json=_build_ingest_payload(observed_at=observed_at, include_scan_metadata=False),
    )

    assert response.status_code == 201

    with client.app.state.session_factory() as session:
        persisted_finding = session.scalar(
            select(models.RawFinding)
            .where(models.RawFinding.run_id == run_id, models.RawFinding.asset_id == "asset-1")
            .order_by(models.RawFinding.finding_id)
        )
        persisted_artifacts = session.scalars(
            select(models.EvidenceArtifact)
            .where(models.EvidenceArtifact.run_id == run_id, models.EvidenceArtifact.asset_id == "asset-1")
            .order_by(models.EvidenceArtifact.artifact_type)
        ).all()

    assert persisted_finding is not None
    assert len(persisted_artifacts) == 3
    assert [artifact.artifact_type.value for artifact in persisted_artifacts] == [
        "dom_snapshot_reference",
        "screenshot",
        "trace",
    ]
    assert {artifact.storage_path for artifact in persisted_artifacts} == {
        "var/evidence/run-1/asset-1/dom.html",
        "var/evidence/run-1/asset-1/failure.png",
        "var/evidence/run-1/asset-1/failure.zip",
    }


def test_get_run_findings_returns_persisted_findings_and_evidence(client: TestClient) -> None:
    create_response = client.post(
        "/runs",
        json={
            "course_url_or_name": "Accessibility Sample Course",
            "auth_metadata": {"method": "placeholder"},
        },
    )
    run_id = create_response.json()["run_id"]
    _seed_asset(client, run_id, "asset-1")
    _persist_findings(client, run_id, "asset-1")

    response = client.get(f"/runs/{run_id}/findings")

    assert response.status_code == 200
    body = response.json()
    assert body["run_id"] == run_id
    assert body["finding_count"] == 2
    assert body["result_counts"] == {
        "violation": 1,
        "pass": 1,
        "incomplete": 0,
        "inapplicable": 0,
    }
    violation = next(item for item in body["findings"] if item["result_type"] == "violation")
    assert violation["asset_id"] == "asset-1"
    assert violation["rule_id"] == "image-alt"
    assert violation["wcag_sc"] == "1.1.1"
    assert [item["artifact_type"] for item in violation["evidence_artifacts"]] == [
        "screenshot",
        "trace",
    ]


def test_get_run_findings_returns_404_for_unknown_run(client: TestClient) -> None:
    response = client.get("/runs/does-not-exist/findings")

    assert response.status_code == 404
    assert response.json() == {"detail": "run 'does-not-exist' does not exist"}
