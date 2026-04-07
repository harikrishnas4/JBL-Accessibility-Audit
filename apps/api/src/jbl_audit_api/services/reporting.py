from __future__ import annotations

import uuid
from collections import Counter, defaultdict
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from jbl_audit_api.core.exceptions import NotFoundError
from jbl_audit_api.db.models import Asset, DefectPriority, RawFinding, ReportRecord, ThirdPartyEvidence
from jbl_audit_api.repositories.reports import ReportRepository

REPORT_TYPE = "excel_export"
REPORT_FILE_NAME = "wcag-audit-report.xlsx"
HOW_TO_FIX_PLACEHOLDER = "HOW TO FIX guidance placeholder until remediation library is implemented."
PRIORITY_ORDER = {
    DefectPriority.P1: 1,
    DefectPriority.P2: 2,
    DefectPriority.P3: 3,
    DefectPriority.P4: 4,
}
PRIORITY_TO_SEVERITY = {
    DefectPriority.P1: "critical",
    DefectPriority.P2: "serious",
    DefectPriority.P3: "moderate",
    DefectPriority.P4: "minor",
}


class LocalReportStorageAdapter:
    def __init__(self, root_directory: Path, uri_prefix: str = "var/reports") -> None:
        self.root_directory = root_directory
        self.uri_prefix = uri_prefix

    def save_workbook(self, run_id: str, workbook: Workbook) -> str:
        run_directory = self.root_directory / run_id
        run_directory.mkdir(parents=True, exist_ok=True)
        file_path = run_directory / REPORT_FILE_NAME
        workbook.save(file_path)
        return f"{self.uri_prefix}/{run_id}/{REPORT_FILE_NAME}"


class ReportingService:
    def __init__(
        self,
        repository: ReportRepository,
        storage_adapter: LocalReportStorageAdapter,
    ) -> None:
        self.repository = repository
        self.storage_adapter = storage_adapter

    def generate_excel_report(self, run_id: str) -> ReportRecord:
        audit_run = self.repository.get_run_context(run_id)
        if audit_run is None:
            raise NotFoundError(f"run '{run_id}' does not exist")

        finding_ids = [
            component.finding_id
            for defect in audit_run.defects
            for component in defect.components
            if component.finding_id
        ]
        raw_findings_by_id = {
            finding.finding_id: finding
            for finding in self.repository.list_raw_findings_by_ids(finding_ids)
        }
        workbook = self._build_workbook(audit_run, raw_findings_by_id)
        report_uri = self.storage_adapter.save_workbook(run_id, workbook)

        now = datetime.now(UTC)
        report_record = self.repository.get_report_record(run_id, REPORT_TYPE)
        if report_record is None:
            report_record = ReportRecord(
                report_record_id=str(uuid.uuid4()),
                run_id=run_id,
                report_type=REPORT_TYPE,
                report_uri=report_uri,
                created_at=now,
                updated_at=now,
            )
        else:
            report_record.report_uri = report_uri
            report_record.updated_at = now
        return self.repository.save_report_record(report_record)

    def _build_workbook(self, audit_run, raw_findings_by_id: dict[str, RawFinding]) -> Workbook:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Audit Summary"
        issues_sheet = workbook.create_sheet("All Issues")
        component_sheet = workbook.create_sheet("Component Health")

        self._populate_summary_sheet(summary_sheet, audit_run, raw_findings_by_id)
        self._populate_all_issues_sheet(issues_sheet, audit_run, raw_findings_by_id)
        self._populate_component_health_sheet(component_sheet, audit_run)
        return workbook

    def _populate_summary_sheet(self, sheet, audit_run, raw_findings_by_id: dict[str, RawFinding]) -> None:
        rows = build_summary_rows(audit_run, raw_findings_by_id)
        sheet.append(["Field", "Value"])
        apply_header_style(sheet, 1)
        for field_name, value in rows:
            sheet.append([field_name, value])
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 40

    def _populate_all_issues_sheet(self, sheet, audit_run, raw_findings_by_id: dict[str, RawFinding]) -> None:
        headers = [
            "issue id",
            "WCAG SC",
            "priority",
            "layer",
            "owner_team",
            "asset URL/locator",
            "asset_type",
            "defect description",
            "HOW TO FIX",
            "evidence path",
            "impacted_asset_count",
            "provider_name",
            "third_party_evidence_status",
            "third_party_evidence_type",
        ]
        sheet.append(headers)
        apply_header_style(sheet, 1)

        assets_by_id = {asset.asset_id: asset for asset in audit_run.assets}
        for defect in sorted(audit_run.defects, key=lambda item: item.issue_id):
            for component in sorted(defect.components, key=lambda item: item.asset_id):
                asset = assets_by_id.get(component.asset_id)
                raw_finding = raw_findings_by_id.get(component.finding_id or "")
                provider_evidence = resolved_third_party_evidence(asset)
                sheet.append(
                    [
                        defect.issue_id,
                        defect.wcag_sc or "",
                        defect.priority.value,
                        defect.layer.value,
                        defect.owner_team or "",
                        asset.locator if asset else component.locator or "",
                        asset.asset_type if asset else "",
                        defect.message,
                        HOW_TO_FIX_PLACEHOLDER,
                        pick_evidence_path(raw_finding),
                        defect.impacted_asset_count,
                        provider_evidence.provider_name if provider_evidence is not None else "",
                        provider_evidence.status if provider_evidence is not None else "",
                        provider_evidence.evidence_type if provider_evidence is not None else "",
                    ],
                )
        set_column_widths(
            sheet,
            {
                "A": 18,
                "B": 12,
                "C": 10,
                "D": 14,
                "E": 16,
                "F": 48,
                "G": 16,
                "H": 48,
                "I": 52,
                "J": 42,
                "K": 18,
                "L": 24,
                "M": 28,
                "N": 26,
            },
        )

    def _populate_component_health_sheet(self, sheet, audit_run) -> None:
        headers = [
            "shared_key",
            "component label",
            "impacted_asset_count",
            "P1",
            "P2",
            "P3",
            "P4",
            "worst severity",
            "owner_team",
            "provider_name",
            "third_party_evidence_status",
            "third_party_evidence_type",
        ]
        sheet.append(headers)
        apply_header_style(sheet, 1)

        for row in build_component_health_rows(audit_run):
            sheet.append(row)

        set_column_widths(
            sheet,
            {
                "A": 28,
                "B": 24,
                "C": 20,
                "D": 8,
                "E": 8,
                "F": 8,
                "G": 8,
                "H": 16,
                "I": 18,
                "J": 24,
                "K": 28,
                "L": 26,
            },
        )


def build_summary_rows(audit_run, raw_findings_by_id: dict[str, RawFinding]) -> list[tuple[str, str | int]]:
    assets = audit_run.assets
    defects = audit_run.defects
    shared_key_counts = Counter(
        resolved_shared_key(asset)
        for asset in assets
        if resolved_shared_key(asset)
    )
    shared_asset_count = sum(
        1
        for asset in assets
        if (key := resolved_shared_key(asset)) and shared_key_counts[key] > 1
    )
    unique_asset_count = len(assets) - shared_asset_count
    defect_priority_counts = Counter(defect.priority.value for defect in defects)
    latest_auth_profile = max(audit_run.auth_profiles, key=lambda item: item.created_at, default=None)
    observed_datetimes = [normalize_datetime(finding.observed_at) for finding in raw_findings_by_id.values()]
    scan_date = max(observed_datetimes) if observed_datetimes else normalize_datetime(audit_run.updated_at)

    return [
        ("course", audit_run.audit_input.course_url_or_name),
        ("run_id", audit_run.run_id),
        ("mode", audit_run.mode.value),
        ("total_assets", len(assets)),
        ("in_scope_assets", sum(1 for asset in assets if asset.scope_status.value == "in_scope")),
        ("out_of_scope_assets", sum(1 for asset in assets if asset.scope_status.value == "out_of_scope")),
        ("shared_assets", shared_asset_count),
        ("unique_assets", unique_asset_count),
        ("defect_count", len(defects)),
        ("P1_defects", defect_priority_counts.get("P1", 0)),
        ("P2_defects", defect_priority_counts.get("P2", 0)),
        ("P3_defects", defect_priority_counts.get("P3", 0)),
        ("P4_defects", defect_priority_counts.get("P4", 0)),
        ("pre_check_status", latest_auth_profile.validation_status.value if latest_auth_profile else "pending"),
        ("scan_date", scan_date.date().isoformat()),
    ]


def build_component_health_rows(audit_run) -> list[list[str | int]]:
    assets_by_group: dict[str, list[Asset]] = defaultdict(list)
    for asset in audit_run.assets:
        assets_by_group[resolved_shared_key(asset) or f"asset:{asset.asset_id}"].append(asset)

    group_priority_counts: dict[str, Counter] = defaultdict(Counter)
    for defect in audit_run.defects:
        linked_groups: set[str] = set()
        for component in defect.components:
            asset = next((item for item in audit_run.assets if item.asset_id == component.asset_id), None)
            if asset is None:
                continue
            linked_groups.add(resolved_shared_key(asset) or f"asset:{asset.asset_id}")
        for group_key in linked_groups:
            group_priority_counts[group_key][defect.priority.value] += 1

    rows: list[list[str | int]] = []
    for group_key in sorted(assets_by_group):
        assets = assets_by_group[group_key]
        priority_counts = group_priority_counts[group_key]
        worst_priority = determine_worst_priority(priority_counts)
        first_asset = assets[0]
        provider_evidence = resolved_third_party_evidence(first_asset)
        rows.append(
            [
                group_key,
                component_label(first_asset),
                len(assets),
                priority_counts.get("P1", 0),
                priority_counts.get("P2", 0),
                priority_counts.get("P3", 0),
                priority_counts.get("P4", 0),
                PRIORITY_TO_SEVERITY[worst_priority] if worst_priority else "",
                resolved_owner_team(first_asset) or "",
                provider_evidence.provider_name if provider_evidence is not None else "",
                provider_evidence.status if provider_evidence is not None else "",
                provider_evidence.evidence_type if provider_evidence is not None else "",
            ],
        )
    return rows


def determine_worst_priority(priority_counts: Counter) -> DefectPriority | None:
    present = [priority for priority in DefectPriority if priority_counts.get(priority.value, 0) > 0]
    if not present:
        return None
    return min(present, key=lambda item: PRIORITY_ORDER[item])


def component_label(asset: Asset) -> str:
    fingerprint = asset.component_fingerprint or {}
    return (
        fingerprint.get("template_id")
        or fingerprint.get("stable_css_selector")
        or asset.asset_type
        or asset.asset_id
    )


def resolved_shared_key(asset: Asset) -> str | None:
    classification = asset.classification_record
    if classification is not None and classification.shared_key:
        return classification.shared_key
    return asset.shared_key


def resolved_owner_team(asset: Asset) -> str | None:
    classification = asset.classification_record
    if classification is not None and classification.owner_team:
        return classification.owner_team
    return asset.owner_team


def resolved_third_party_evidence(asset: Asset | None) -> ThirdPartyEvidence | None:
    if asset is None or asset.classification_record is None:
        return None
    return asset.classification_record.third_party_evidence


def pick_evidence_path(raw_finding: RawFinding | None) -> str:
    if raw_finding is None:
        return ""
    ranked_artifacts = sorted(
        raw_finding.evidence_artifacts,
        key=lambda artifact: artifact_rank(artifact.artifact_type.value),
    )
    return ranked_artifacts[0].storage_path if ranked_artifacts else ""


def artifact_rank(artifact_type: str) -> int:
    return {
        "screenshot": 0,
        "trace": 1,
        "dom_snapshot_reference": 2,
    }.get(artifact_type, 99)


def apply_header_style(sheet, row_index: int) -> None:
    for cell in sheet[row_index]:
        cell.font = Font(bold=True)


def set_column_widths(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def normalize_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=UTC)
    return value.astimezone(UTC)
