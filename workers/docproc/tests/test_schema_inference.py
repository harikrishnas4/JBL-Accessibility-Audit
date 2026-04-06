from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from jbl_docproc.schema_inference.engine import SchemaInferenceEngine
from jbl_docproc.schema_inference.models import CanonicalSchemaType, ConfidenceTier, InferenceReport, WorkbookInventory
from jbl_docproc.schema_inference.registry import SchemaRegistryRepository


class FakeRegistryRepository(SchemaRegistryRepository):
    def __init__(self) -> None:
        self.lookup_called = False
        self.saved_report: InferenceReport | None = None

    def lookup(self, workbook_inventory: WorkbookInventory):
        self.lookup_called = True
        return ()

    def save(self, report: InferenceReport) -> None:
        self.saved_report = report


def test_known_workbook_maps_all_canonical_schema_types(tmp_path: Path) -> None:
    workbook_path = tmp_path / "known_manifest.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_sheet(
        workbook,
        "chapter_toc",
        [
            ["chapter_id", "chapter_title", "chapter_order"],
            ["CH-01", "Introduction", 1],
            ["CH-02", "Advanced Concepts", 2],
        ],
    )
    _add_sheet(
        workbook,
        "asset_type_layout",
        [
            ["asset_type", "layout", "template"],
            ["video", "grid", "video_grid"],
            ["article", "two column", "article_two_col"],
        ],
    )
    _add_sheet(
        workbook,
        "topic_ordering",
        [
            ["topic_id", "topic_name", "order"],
            ["TOP-001", "Foundations", 1],
            ["TOP-002", "Review", 2],
        ],
    )
    _add_sheet(
        workbook,
        "embed_registry",
        [
            ["embed_id", "embed_type", "embed_url"],
            ["EMB-01", "youtube", "https://youtube.com/watch?v=abc"],
        ],
    )
    _add_sheet(
        workbook,
        "label_map",
        [
            ["source_label", "canonical_label", "locale"],
            ["hero image", "hero_image", "en-US"],
        ],
    )
    _add_sheet(
        workbook,
        "document_url_map",
        [
            ["document_id", "document_title", "document_url"],
            ["DOC-1", "Guide", "https://cdn.example.com/guide.pdf"],
        ],
    )
    _add_sheet(
        workbook,
        "media_categories",
        [
            ["media_id", "media_type", "category"],
            ["MED-001", "video", "core"],
        ],
    )
    workbook.save(workbook_path)

    registry = FakeRegistryRepository()
    schema_map, report = SchemaInferenceEngine(registry_repository=registry).infer(workbook_path)

    assert report.workbook_inventory.sheet_names == (
        "chapter_toc",
        "asset_type_layout",
        "topic_ordering",
        "embed_registry",
        "label_map",
        "document_url_map",
        "media_categories",
    )
    assert len(report.workbook_inventory.sheets) == 7
    chapter_inventory = report.workbook_inventory.sheets[0]
    assert chapter_inventory.row_count == 3
    assert chapter_inventory.column_count == 3
    assert chapter_inventory.header_candidates[0].normalized_values == (
        "chapter id",
        "chapter title",
        "chapter order",
    )
    assert chapter_inventory.sample_rows[0].values == ("CH-01", "Introduction", "1")
    assert registry.lookup_called is True
    assert registry.saved_report == report

    for schema_type in CanonicalSchemaType:
        assignment = schema_map.get(schema_type)
        assert assignment is not None
        assert assignment.confidence == ConfidenceTier.high


def test_partial_workbook_returns_subset_with_medium_or_better_confidence(tmp_path: Path) -> None:
    workbook_path = tmp_path / "partial_manifest.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "topic sequence"
    sheet.append(["topic name", "sort order", "chapter"])
    sheet.append(["Keyboard Basics", 1, "CH-01"])
    sheet.append(["Screen Readers", 2, "CH-01"])
    workbook.save(workbook_path)

    schema_map, report = SchemaInferenceEngine().infer(workbook_path)

    assert report.workbook_inventory.sheet_names == ("topic sequence",)
    assignment = schema_map.get(CanonicalSchemaType.topic_ordering)
    assert assignment is not None
    assert assignment.confidence in {ConfidenceTier.medium, ConfidenceTier.high}
    assert schema_map.unmatched_sheets == ()


def test_low_confidence_workbook_is_flagged_low(tmp_path: Path) -> None:
    workbook_path = tmp_path / "weak_manifest.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "chapter overview"
    sheet.append(["chapter title", "notes"])
    sheet.append(["Accessibility Intro", "starter"])
    workbook.save(workbook_path)

    schema_map, _ = SchemaInferenceEngine().infer(workbook_path)

    assignment = schema_map.get(CanonicalSchemaType.chapter_toc)
    assert assignment is not None
    assert assignment.confidence == ConfidenceTier.low


def test_no_manifest_case_returns_no_assignments(tmp_path: Path) -> None:
    workbook_path = tmp_path / "no_manifest.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "notes"
    sheet.append(["comment", "owner"])
    sheet.append(["needs review", "team"])
    workbook.save(workbook_path)

    schema_map, report = SchemaInferenceEngine().infer(workbook_path)

    assert schema_map.assignments == ()
    assert schema_map.unmatched_sheets == ("notes",)
    assert report.sheet_inferences[0].confidence == ConfidenceTier.none


def _add_sheet(workbook: Workbook, title: str, rows: list[list[object]]) -> None:
    sheet = workbook.create_sheet(title=title)
    for row in rows:
        sheet.append(row)
