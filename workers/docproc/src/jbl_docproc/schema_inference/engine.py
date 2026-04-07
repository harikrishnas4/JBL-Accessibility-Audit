from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from rapidfuzz import fuzz

from jbl_docproc.schema_inference.definitions import SCHEMA_DEFINITIONS, SchemaDefinition
from jbl_docproc.schema_inference.models import (
    ConfidenceTier,
    HeaderCandidate,
    InferenceReport,
    SampleRow,
    SchemaAssignment,
    SchemaMap,
    ScoreBreakdown,
    SheetInference,
    SheetInventory,
    WorkbookInventory,
)
from jbl_docproc.schema_inference.registry import NullSchemaRegistryRepository, SchemaRegistryRepository


class SchemaInferenceEngine:
    def __init__(self, registry_repository: SchemaRegistryRepository | None = None) -> None:
        self.registry_repository = registry_repository or NullSchemaRegistryRepository()

    def inventory_workbook(self, workbook_path: str | Path) -> WorkbookInventory:
        resolved_path = Path(workbook_path).resolve()
        workbook = load_workbook(filename=resolved_path, read_only=True, data_only=True)
        try:
            sheet_inventories = tuple(self._inventory_sheet(workbook[sheet_name]) for sheet_name in workbook.sheetnames)
            return WorkbookInventory(
                workbook_path=resolved_path,
                sheet_names=tuple(workbook.sheetnames),
                sheets=sheet_inventories,
            )
        finally:
            workbook.close()

    def infer(self, workbook_path: str | Path) -> tuple[SchemaMap, InferenceReport]:
        workbook_inventory = self.inventory_workbook(workbook_path)
        registry_records = tuple(self.registry_repository.lookup(workbook_inventory))

        sheet_reports: list[SheetInference] = []
        accepted_assignments: list[SchemaAssignment] = []
        matched_sheet_names: set[str] = set()

        for sheet_inventory in workbook_inventory.sheets:
            best_assignment = self._best_assignment(sheet_inventory)
            sheet_reports.append(
                SheetInference(
                    sheet_name=sheet_inventory.sheet_name,
                    schema_type=(
                        best_assignment.schema_type
                        if best_assignment.confidence != ConfidenceTier.none
                        else None
                    ),
                    confidence=best_assignment.confidence,
                    matched_header_row=best_assignment.matched_header_row,
                    matched_headers=best_assignment.matched_headers,
                    pattern_hits=best_assignment.pattern_hits,
                    score_breakdown=best_assignment.score_breakdown,
                ),
            )
            if best_assignment.confidence != ConfidenceTier.none:
                accepted_assignments.append(best_assignment)
                matched_sheet_names.add(sheet_inventory.sheet_name)

        schema_map = SchemaMap(
            assignments=tuple(sorted(accepted_assignments, key=lambda item: item.schema_type.value)),
            unmatched_sheets=tuple(
                sheet_name for sheet_name in workbook_inventory.sheet_names if sheet_name not in matched_sheet_names
            ),
        )
        report = InferenceReport(
            workbook_inventory=workbook_inventory,
            schema_map=schema_map,
            sheet_inferences=tuple(sheet_reports),
            registry_records=registry_records,
        )
        self.registry_repository.save(report)
        return schema_map, report

    def _inventory_sheet(self, worksheet: Any) -> SheetInventory:
        row_limit = min(int(worksheet.max_row or 0), 8)
        rows = list(worksheet.iter_rows(min_row=1, max_row=max(row_limit, 1), values_only=True))
        header_candidates = self._extract_header_candidates(rows)
        sample_rows = self._extract_sample_rows(rows, header_candidates)
        return SheetInventory(
            sheet_name=worksheet.title,
            row_count=int(worksheet.max_row or 0),
            column_count=int(worksheet.max_column or 0),
            header_candidates=tuple(header_candidates),
            sample_rows=tuple(sample_rows),
        )

    def _extract_header_candidates(self, rows: list[tuple[Any, ...]]) -> list[HeaderCandidate]:
        candidates: list[HeaderCandidate] = []
        for index, row in enumerate(rows[:5], start=1):
            raw_values = tuple(self._stringify(value) for value in row if self._stringify(value))
            if len(raw_values) < 2:
                continue
            normalized_values = tuple(self._normalize_token(value) for value in raw_values)
            if len(set(normalized_values)) < 2:
                continue
            candidates.append(
                HeaderCandidate(
                    row_index=index,
                    raw_values=raw_values,
                    normalized_values=normalized_values,
                ),
            )
        return candidates

    def _extract_sample_rows(
        self,
        rows: list[tuple[Any, ...]],
        header_candidates: list[HeaderCandidate],
    ) -> list[SampleRow]:
        start_row = header_candidates[0].row_index + 1 if header_candidates else 1
        samples: list[SampleRow] = []
        for index, row in enumerate(rows[start_row - 1 :], start=start_row):
            values = tuple(self._stringify(value) for value in row if self._stringify(value))
            if not values:
                continue
            samples.append(SampleRow(row_index=index, values=values))
            if len(samples) == 3:
                break
        return samples

    def _best_assignment(self, sheet_inventory: SheetInventory) -> SchemaAssignment:
        assignments = [self._score_sheet(sheet_inventory, definition) for definition in SCHEMA_DEFINITIONS]
        assignments.sort(key=lambda item: item.score_breakdown.total_score, reverse=True)
        return assignments[0]

    def _score_sheet(self, sheet_inventory: SheetInventory, definition: SchemaDefinition) -> SchemaAssignment:
        sheet_name_score = max(
            fuzz.token_sort_ratio(
                self._normalize_token(sheet_inventory.sheet_name),
                self._normalize_token(alias),
            )
            / 100.0
            for alias in definition.sheet_name_aliases
        )

        header_score, matched_header_row, matched_headers = self._header_overlap_score(sheet_inventory, definition)
        pattern_score, pattern_hits = self._pattern_score(sheet_inventory, definition)

        total_score = round((sheet_name_score * 0.4) + (header_score * 0.4) + (pattern_score * 0.2), 4)
        return SchemaAssignment(
            schema_type=definition.schema_type,
            sheet_name=sheet_inventory.sheet_name,
            confidence=self._confidence_tier(total_score),
            matched_header_row=matched_header_row,
            matched_headers=matched_headers,
            pattern_hits=pattern_hits,
            score_breakdown=ScoreBreakdown(
                sheet_name_score=round(sheet_name_score, 4),
                header_overlap_score=round(header_score, 4),
                data_pattern_score=round(pattern_score, 4),
                total_score=total_score,
            ),
        )

    def _header_overlap_score(
        self,
        sheet_inventory: SheetInventory,
        definition: SchemaDefinition,
    ) -> tuple[float, int | None, tuple[str, ...]]:
        best_score = 0.0
        best_row: int | None = None
        best_matches: tuple[str, ...] = ()

        for candidate in sheet_inventory.header_candidates:
            field_scores: list[float] = []
            matched_fields: list[str] = []
            for field_name, aliases in definition.header_aliases.items():
                score = self._best_header_match(candidate.normalized_values, aliases)
                field_scores.append(score)
                if score >= 0.75:
                    matched_fields.append(field_name)
            candidate_score = sum(field_scores) / len(field_scores) if field_scores else 0.0
            if candidate_score > best_score:
                best_score = candidate_score
                best_row = candidate.row_index
                best_matches = tuple(matched_fields)

        return best_score, best_row, best_matches

    def _best_header_match(self, headers: tuple[str, ...], aliases: tuple[str, ...]) -> float:
        best = 0.0
        normalized_aliases = tuple(self._normalize_token(alias) for alias in aliases)
        for header in headers:
            for alias in normalized_aliases:
                score = fuzz.ratio(header, alias) / 100.0
                if score > best:
                    best = score
        return best if best >= 0.65 else 0.0

    def _pattern_score(
        self,
        sheet_inventory: SheetInventory,
        definition: SchemaDefinition,
    ) -> tuple[float, tuple[str, ...]]:
        if not definition.data_patterns:
            return 0.0, ()
        flattened_values = [value for row in sheet_inventory.sample_rows for value in row.values]
        hits = tuple(
            pattern_name
            for pattern_name, regex in definition.data_patterns.items()
            if self._has_pattern_hit(flattened_values, regex)
        )
        return len(hits) / len(definition.data_patterns), hits

    def _has_pattern_hit(self, values: list[str], regex: re.Pattern[str]) -> bool:
        return any(regex.search(value) for value in values)

    def _confidence_tier(self, score: float) -> ConfidenceTier:
        if score >= 0.8:
            return ConfidenceTier.high
        if score >= 0.6:
            return ConfidenceTier.medium
        if score >= 0.35:
            return ConfidenceTier.low
        return ConfidenceTier.none

    def _stringify(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _normalize_token(self, value: str) -> str:
        return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()
