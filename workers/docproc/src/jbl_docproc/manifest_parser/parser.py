from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from rapidfuzz import fuzz

from jbl_docproc.schema_inference.definitions import SCHEMA_DEFINITIONS
from jbl_docproc.schema_inference.models import CanonicalSchemaType, SchemaMap

from .models import CanonicalDataset, CanonicalRecord, ManifestParseResult


class ManifestParser:
    def parse(self, workbook_path: str | Path, schema_map: SchemaMap) -> ManifestParseResult:
        resolved_path = Path(workbook_path).resolve()
        workbook = load_workbook(filename=resolved_path, read_only=True, data_only=True)
        try:
            datasets: list[CanonicalDataset] = []
            for assignment in schema_map.assignments:
                worksheet = workbook[assignment.sheet_name]
                definition = next(
                    definition
                    for definition in SCHEMA_DEFINITIONS
                    if definition.schema_type == assignment.schema_type
                )
                rows = list(worksheet.iter_rows(values_only=True))
                header_row_index = assignment.matched_header_row or 1
                header_values = rows[header_row_index - 1] if rows and len(rows) >= header_row_index else ()
                column_indexes, column_mapping = self._resolve_columns(header_values, definition.header_aliases)
                records = self._extract_records(rows, header_row_index, column_indexes)
                datasets.append(
                    CanonicalDataset(
                        schema_type=assignment.schema_type,
                        sheet_name=assignment.sheet_name,
                        matched_header_row=assignment.matched_header_row,
                        column_mapping=column_mapping,
                        records=tuple(records),
                    ),
                )

            found = {dataset.schema_type for dataset in datasets}
            fallback_flags = tuple(
                f"missing_{schema_type.value}"
                for schema_type in CanonicalSchemaType
                if schema_type not in found
            )
            return ManifestParseResult(
                datasets=tuple(sorted(datasets, key=lambda dataset: dataset.schema_type.value)),
                fallback_flags=fallback_flags,
            )
        finally:
            workbook.close()

    def _resolve_columns(
        self,
        header_values: tuple[Any, ...],
        header_aliases: dict[str, tuple[str, ...]],
    ) -> tuple[dict[str, int], dict[str, str]]:
        normalized_headers = [self._normalize(self._stringify(value)) for value in header_values]
        resolved_indexes: dict[str, int] = {}
        resolved_headers: dict[str, str] = {}
        used_indexes: set[int] = set()

        for field_name, aliases in header_aliases.items():
            best_index = None
            best_score = 0.0
            for index, header in enumerate(normalized_headers):
                if index in used_indexes or not header:
                    continue
                for alias in aliases:
                    score = fuzz.ratio(header, self._normalize(alias)) / 100.0
                    if score > best_score:
                        best_score = score
                        best_index = index
            if best_index is not None and best_score >= 0.65:
                used_indexes.add(best_index)
                resolved_indexes[field_name] = best_index
                resolved_headers[field_name] = self._stringify(header_values[best_index])
        return resolved_indexes, resolved_headers

    def _extract_records(
        self,
        rows: list[tuple[Any, ...]],
        header_row_index: int,
        column_indexes: dict[str, int],
    ) -> list[CanonicalRecord]:
        records: list[CanonicalRecord] = []
        if not column_indexes:
            return records

        for row_index, row in enumerate(rows[header_row_index:], start=header_row_index + 1):
            values = {
                field_name: self._stringify(row[column_index])
                for field_name, column_index in column_indexes.items()
                if column_index < len(row) and self._stringify(row[column_index])
            }
            if values:
                records.append(CanonicalRecord(row_index=row_index, values=values))
        return records

    def _stringify(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _normalize(self, value: str) -> str:
        return " ".join(value.lower().replace("_", " ").split())
